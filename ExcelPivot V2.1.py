# Read csv exported from VoterLetters and report
# V1.1 produce a tab for each team
# V1.2 remove all the first wee for grouping - use offset AND input modified file containing multi-org in one
#     also removed ExcelWriter section which wasn't used
# V1/3 Runs off Aarons file across all orgs & campaigns.  Puts reports in dir based off input file name.
#   Some vars were renamed in Aaron's file.
# V1.4 report by month as well as week
# V 1.5 added pivot reports for ROV enterprise across all orgs
#     use fulfilled_count rather than requested_count in tables
# V1.6 incorporate autosize_xls_cols to size cols before titles added
#  1/25/22 Split Team reports into Summ and wWriters
# V2.0 incorporate uploading to google sheets and adding permissions.  Started as ExcelPivot V1.6
# Updates taken from "VL reports via Google API V1.0"
# Finding and creating a folder for upload taken from "find folder google api.py" and "create sub folder with Google API.py"
# V2.1 Uploaded the single Admin report

# TODO: Move off my personal google account
# TODO: Add state tab before campaigns to show state totals.
# TODO: Add a tab or separate report that shows writers who have never requested.   How to avoide those invite to receive only and not order?



from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from datetime import datetime
import datetime as dt
import xlsxwriter
import pymsgbox
import os
from pathlib import *
import pathlib
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from random import random
import numpy as np
import pandas as pd
from tkinter import *
from tkinter import Tk  # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory
from tkinter import messagebox
import ast
import google.auth.transport.requests
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from apiclient.http import MediaFileUpload
# from apiclient.discovery import build

# for google drive api
SCOPES = ['https://www.googleapis.com/auth/drive']
ORG_REPORT_FOLDER_ID = "1OEAdTnhQoAKpyzFqYdM1ztr42vHAqbk5"  # Folder holding VoterLetters Organizer Reports
ADMIN_REPORT_FOLDER_ID ='1_vTXsMYOwK_TLYaRqC7YA_jolBuqDWP4'

sendPermissionAddEmail = True  # send permission granted emails


noteLines = [
    '',
    'WHAT IS IN THIS REPORT',
    '',
    'What you’ll get:',
    '   - Each spreadsheet presents the address and request counts in your room: by time-period,'
    'state, campaign, team, and writer.',
    '   - The tabs (located at the bottom of the sheet) provide an overview then gets more specific:',
    '       the first few tabs show data for your whole room.  Tabs after that are one for each team.',
    '   - A weekly version showing counts by week will be produced each Monday. Monthly time periods',
    '       will be shown in a report produced on the first day of each month.',
    '',
    'The overall report:',
    '   - The sheet shows only the address and request counts for your room.',
    '   - Each tab shows totals over time, either in weekly or monthly increments',
    '   - The titles give the',
    '         - name of your room',
    '         - name of the team',
    '         - time increment (weekly or monthly)',
    '         - dates included in the report',
    '   - The data will be summarized in each row',
    '   - The reports show total state, campaign, teams, and writers',
    '',
    'The individual tabs of the report are:',
    '   - Tab 1, named  “Notes” explains the report'
    '   - Tab 2, named  “Campaigns” shows requested addresses broken by Campaign',
    '   - Tab 3, “All Team Sums” by Teams',
    '   - Tab 4, “Teams w Writers” by Writers within Teams',
    '   - Tab 5, “Teams w Counts” number of requests by Writers within Teams. ',
    '   - The tabs following are specific to each team (including individual writers).',
    '       These can be copied and pasted for team leads.  '
]

def autosizexls(ws):
    # from https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    # print(ws.title)
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                # print("cell value", cell.value, "data_type",cell.data_type)
                if(cell.data_type == 'd'):
                    dateWidth = 10
                else: dateWidth = len(str(cell.value))
                # dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), dateWidth))

    for col, value in dims.items():
        # print(col , value)
        ws.column_dimensions[col].width = value + 1

def getFolderIds(foldernm, par):
    q1 = "mimeType='application/vnd.google-apps.folder' and trashed=false and name='" + foldernm + "' and parents = '" + par +"'"
    idList = []
    page_token = None
    while True:
        response = drive_service.files().list(
            q=q1,
            spaces='drive',
            fields="nextPageToken, files(id, name, trashed, parents)",
            pageToken=page_token).execute()
        print(response)
        for file in response.get('files', []):
            idList.append(file.get('id'))
            print('Found folder (name,id,trashed,parent):', (file.get('name'), file.get('id'), file.get('trashed'), file.get('parents')))

        page_token = response.get('nextPageToken', None)
        if page_token is None:
            break
    return idList

def getFileIds(filenm, par):
    q1 = "mimeType != 'application/vnd.google-apps.folder' and trashed=false and name='" + filenm + "' and parents = '" + par +"'"
    idList = []
    page_token = None
    while True:
        response = drive_service.files().list(
            q=q1,
            spaces='drive',
            fields="nextPageToken, files(id, name, trashed, parents)",
            pageToken=page_token).execute()
        print(response)
        for file in response.get('files', []):
            idList.append(file.get('id'))
            print('Found file (name,id,trashed,parent):', (file.get('name'), file.get('id'), file.get('trashed'), file.get('parents')))

        page_token = response.get('nextPageToken', None)
        if page_token is None:
            break
    return idList

def getCreds():
    # get credentials needed to set up API service.  Was inline code; this cleaned things up.
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return creds

#### Start Main code
####
# Identify which VoterLetters files should be downloaded before starting
pymsgbox.alert("Download data from VoterLetters before running.\n\n"\
               "1. Addresses assigned:\n      VoterLetters ROV >\n      Reports >\n      New Report >\n      All Parent Campaign Address Requests.\n  Dates 1/1/22 to prior Monday INCLUSIVE.\n\n"\
               "2. Users (to assign google read permissions):\n      VoterLetters Child Organizations >\n      Export Users to CSV.\n",\
               "Get ready!")

choice = pymsgbox.confirm("Run Reports or upload files to Google Drive", 'Run, Upload or Exit?',
                          ["Run", 'Upload', 'Exit'])
if choice == "Exit":
    pymsgbox.alert("Exiting", "Alert")
    exit()
elif choice == "Upload":
    # pymsgbox.alert("** Uploading Here  **", "Alert")
    if sendPermissionAddEmail:
        choice = pymsgbox.confirm("Currently sending google permission emails.", 'Send emails?',
                                  ["OK", 'Cancel'])
        if choice == "Cancel":
            # pymsgbox.alert("Exiting", "Alert")
            exit()
    if True:  # use TK input file, set file name, or small test list
        if True:
            Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
            inputDir = os.path.expanduser("/Users/Denise/Downloads/")
            InputFile = askopenfilename(initialdir=os.path.expanduser(inputDir),
                        title="Select VoterLetters export file for USER LIST (enterprise-users-yyyy-mm-dd.csv)",
                        filetypes=(("CSV files",
                                    "*.csv "),))  # show an "Open" dialog box and return the path to the selected file
            if InputFile == "":
                pymsgbox.alert("No file chosen - exiting", "Alert")
                exit()
            df = pd.read_csv(InputFile,
                             usecols=['name', 'email', 'role', 'organization'])
        else:
            df = pd.read_csv("/Users/Denise/Downloads/enterprise-users-2022-01-21.csv",
                             usecols=['name', 'email', 'role', 'organization'])
        df = df.sort_values(["organization", "name"], ascending=(True, True))
        orgEmailList = df.loc[df['role'].isin(['organizer']), ["organization", "email"]].values.tolist()
    else:
        orgEmailList = [
            ['NY-NYC and Long Island', 'kramsman+nycorg@gmail.com'],
            ['NY-NYC and Long Island', 'meika@centerforcommonground.org'],
            ['CA-Peninsula-South Bay', 'dee@centerforcommonground.org'],
            ['Mountain States', 'comstockrov@gmail.com'],
            ['CA-Peninsula-South Bay', 'nancy@centerforcommonground.org'],
        ]

    creds = getCreds() # function call to clean up

    try:
        drive_service = build('drive', 'v3',
                credentials=creds)  # copied from https://stackoverflow.com/questions/11472401/looking-for-example-using-mediafileupload

        Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
        dirStartLoc = os.path.expanduser("~/Dropbox/Postcard Files/VL Org Reports")
        strReportDirToUpload = askdirectory(initialdir=dirStartLoc, title="Select report directory to UPLOAD")
        # strReportDirToUpload = '/Users/Denise/Dropbox/Postcard Files/VL Org Reports/all-parent-campaigns-requests-2022-01-26-W'
        if strReportDirToUpload == "":
            pymsgbox.alert("No directory chosen - exiting", "Alert")
            exit()
        reportDirToUpload = pathlib.Path(strReportDirToUpload)
        folderName = reportDirToUpload.parts[-1]  # the last section of the path
        reportBy = strReportDirToUpload[-1]

        # id list of folders to trash(delete)
        deleteList = getFolderIds(folderName, ORG_REPORT_FOLDER_ID)
        for folderId in deleteList:
            drive_service.files().update(fileId=folderId, body={'trashed':True}).execute()  # trash instead of delete moves to trash; delete gone for good
        a=1

        # Create sub folder
        file_metadata = {
                'name': folderName,
                'parents': [ORG_REPORT_FOLDER_ID],
                'mimeType': 'application/vnd.google-apps.folder' }
        file = drive_service.files().create(body=file_metadata,
                fields='id').execute()
        newFolderId = file.get('id')
        print('Folder ID created: %s' % file.get('id'))

        # get list of filenames in computer directory
        excelReportFiles = glob.glob(str(reportDirToUpload / '[!~]*.xlsx'))  # xlsx files not starting with ~ (temporary files)
        for fileWPath in excelReportFiles:
            # filePath = pathlib.Path('/Users/Denise/Dropbox/Postcard Files/VoterLetters Reports/all-parent-campaigns-requests-2022-01-20-W')
            # path, fileNameWOPath = os.path.split(fileWPath)
            fileName = os.path.basename(fileWPath)  # returns filename with extension
            fileNameWOExt = pathlib.Path(fileWPath).stem  # filename without extension
            roomName = fileNameWOExt[0:-34]  # lop off the additional 34 chars of info added to room for filename like date
            print('Uploading filename, RoomName ', fileName, roomName)

            file_metadata = {'name': fileNameWOExt,
                    'mimeType': 'application/vnd.google-apps.spreadsheet',
                    "parents": [newFolderId]}
            media = MediaFileUpload(fileWPath,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')  # converts xlsx
            file = drive_service.files().create(body=file_metadata, supportsAllDrives=True,
                    media_body=media,
                    fields='id').execute()
            print('Uploaded file ID: %s' % file.get('id'))
            uploadedFileId = file.get('id')

            for org, email in orgEmailList:
                # email = 'briank@kramericore.com'
                if org == roomName:
                    print('ready to add permission for email (org, roomName, email): ', org, roomName, email)
                    payload = {
                        "role": "reader",
                        "type": "user",
                        "emailAddress": email
                    }
                    if sendPermissionAddEmail:
                        if fileNameWOExt[-2:].lower() == '-m':
                            emailMessageMsg = "A new MONTHLY VoterLetters summary report is available for your room.  Click to open."
                        else:
                            emailMessageMsg = "A new WEEKLY VoterLetters summary report is available for your room.  Click to open."
                        drive_service.permissions().create(fileId= uploadedFileId, sendNotificationEmail = True, emailMessage = emailMessageMsg, body=payload).execute()
                    else:
                        # Same as above .create command but do not include email text as not allowed if no email sent
                        drive_service.permissions().create(fileId= uploadedFileId, sendNotificationEmail = False, body=payload).execute()

        # Upload single Admin report
        Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
        dirStartLoc = os.path.expanduser("~/Dropbox/Postcard Files/VL Admin Reports")
        strAdminReportToUpload = askopenfilename(initialdir=dirStartLoc,
                title="Select ADMIN report to UPLOAD",
                filetypes=(("Xlsx files", "*.xlsx "),))
        if strAdminReportToUpload != "":
            # do the upload
            adminReportToUpload = pathlib.Path(strAdminReportToUpload)
            fileName = os.path.basename(adminReportToUpload)  # returns filename with extension
            fileNameWOExt = pathlib.Path(adminReportToUpload).stem  # filename without extension

            deleteList = getFileIds(fileNameWOExt, ADMIN_REPORT_FOLDER_ID)
            for fileId in deleteList:
                drive_service.files().update(fileId=fileId, body={'trashed':True}).execute()  # trash instead of delete moves to trash; delete gone for good
            print('Uploading filename ', fileNameWOExt)

            file_metadata = {'name': fileNameWOExt,
                    'mimeType': 'application/vnd.google-apps.spreadsheet',
                    "parents": [ADMIN_REPORT_FOLDER_ID]}
            media = MediaFileUpload(adminReportToUpload,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')  # converts xlsx
            file = drive_service.files().create(body=file_metadata, supportsAllDrives=True,
                    media_body=media,
                    fields='id').execute()
            print('Uploaded file ID: %s' % file.get('id'))

    except HttpError as err:
        print(err)

elif choice == "Run":
    # pymsgbox.alert("** Running Here  **", "Alert")

    reportBy = pymsgbox.confirm('[W]eekly or [M]onthly report?', 'Date Format', ["W", "M", 'Cancel'])
    if reportBy.lower() == 'm':
        reportVar = "month"
    elif reportBy.lower() == 'w':
        reportVar = "data_week_of"
    else:
        exit()

    # Put enterprise wide reports in a separate directory
    # outputDirAdmin = os.path.expanduser("~/Dropbox/Postcard Files/Other/VoterLetters/Reports/VL ROVWide Reports")
    outputDirAdmin = os.path.expanduser("~/Dropbox/Postcard Files/VL Admin Reports")

    # Dir to prompt to start looking for input file
    inputDir = os.path.expanduser("/Users/Denise/Downloads/")
    # Base of report file; input file name is used to create a sub dir under this
    outputDir = os.path.expanduser("~/Dropbox/Postcard Files/VL Org Reports")

    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing

    # InputFile = askopenfilename(initialdir=os.path.expanduser(r"/Users/Denise/Downloads/"),title="Select VoterLetters address export file", filetypes=(("CSV files", "*.csv "),))  # show an "Open" dialog box and return the path to the selected file
    InputFile = askopenfilename(initialdir=os.path.expanduser(inputDir),title="Select VoterLetters address export file (all-parent-campaigns-requests-yyyy-mm-dd.csv)", filetypes=(("CSV files", "*.csv "),))  # show an "Open" dialog box and return the path to the selected file
    # tried but failed InputFile = askopenfilename(initialdir=os.path.expanduser(inputDir),title="Select VoterLetters address export file", filetypes=('VL files','all-parent*.*'))  # show an "Open" dialog box and return the path to the selected file
    # TODO: set filter to only display fileNAMES that are proper input, such as 'all-parent*.csv'

    # head, tail = os.path.split(InputFile)
    VLdatafile = os.path.basename(InputFile)
    outputDirWFile = os.path.join(outputDir,Path(InputFile).stem + "-" + reportBy)
    filedate = Path(InputFile).stem[-10:]

    # CHECK and create dir if not exists
    if not os.path.isdir(outputDirWFile):
        # pymsgbox.alert("Input Path does not exist. Creating.\n\n" + outputDirWFile, "Fixing Directory Structure")
        # FIXME: close Tkinter window here
        os.makedirs(outputDirWFile)

    # Create dataframe of excel sheet data
    xl = pd.read_csv(InputFile)

    # replace '/' in some input fields (like org name) which cause issues when used as directories
    xl['org_name'] = xl['org_name'].str.replace("/","-").str.replace("'","-").str.replace(",","-")
    xl['team_name'] = xl['team_name'].str.replace("/","-").str.replace("'","-").str.replace(",","-")

    # create a date field as a datetime object
    xl['date2'] = pd.to_datetime(xl['created_at'])
    minDate2 = xl['created_at'].min()
    maxDate2 = xl['created_at'].max()
    # xl['date2'] = pd.to_datetime("2021/02/01")

    # 'daysoffset' will contain the weekday 'day of week' as integers so we can step back to Monday
    xl['daysoffset'] = xl['date2'].apply(lambda x: x.weekday())
    # We apply, row by row (axis=1) a timedelta operation
    xl['data_week_of'] = xl.apply(lambda x: x['date2'] - dt.timedelta(days=x['daysoffset']), axis=1).dt.date
    xl['month'] = pd.to_datetime(xl['date2']).dt.to_period('M')

    xl['team2']=xl['team_name'] # Need two copies of team variable for pivots - one for row, one for cols # FIXME: I believe this was removed and two tables used
    xl.fillna(" No Team", inplace = True) # note space in front for sorting
    xl = xl[~xl['org_name'].isin(['ROV Test Silo','ROV Training Silo','ROV Sample Silo'])]  # select records for one Org
    teams = xl['team_name'].unique()
    teams.sort()

    state_list = ["AL", "AK", "AS", "AZ", "AR", "CA", "CO", "CT", "DE", "DC", "FL", "GA", "GU", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "MP", "OH", "OK", "OR", "PA", "PR", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "VI", "WA", "WV", "WI", "WY"]
    xl['master_campaign'] = xl['parent_campaign_name'].apply(lambda x: x[0:2])
    master_campaigns = xl['master_campaign'].unique()
    # delete those that are not states (some campaign names ended up 'do not use VA-Loudon...' etc  # FIXME: we may want to include these even though state is bad; many werre written in "DO not use VA-xxx"
    master_campaigns = [x for x in master_campaigns if x in state_list]
    master_campaigns.sort()

    ### Create a sheet w pivot tabs for all orgs in separate directory
    print("Org: Enterprise wide")

    file = os.path.join(os.path.expanduser(outputDirAdmin), "ROVWide VoterLetters Summary " + filedate + "-" + reportBy + ".xlsx")

    # Use Pandas dataframes here because it supports pivot tables and openpyxl does not
    # now use writer because it works with pandas dataframes
    writer = pd.ExcelWriter(file)

    # Run standalone pivot on campaigns
    # df_pt = pd.pivot_table(xlo, columns=['data_week_of'], index=['parent_campaign_name', 'team_name'], values=['addresses_count'], margins=True, dropna=True,aggfunc=np.sum)
    df_pt = pd.pivot_table(xl, columns=[reportVar], index=['master_campaign', 'parent_campaign_name'],
                           values=['addresses_count'], margins=True, dropna=True, aggfunc=np.sum)
    df_pt.to_excel(writer, sheet_name='Campaigns', startrow=6)  # summary sheet of al teams combined

    df_pt = pd.pivot_table(xl, columns=[reportVar], index=['org_name'],
                           values=['addresses_count'], margins=True, dropna=True, aggfunc=np.sum)
    df_pt.to_excel(writer, sheet_name='Rooms', startrow=6)  # summary sheet of al teams combined

    for m in master_campaigns:
        xlm = xl[xl['master_campaign'] == m] # select records for one Org
        df_pt = pd.pivot_table(xlm, columns=[reportVar], index=['master_campaign', 'parent_campaign_name'], values=['addresses_count'],
                               margins=True, dropna=True, aggfunc=np.sum)
        # print("Sheet name = ", m)
        df_pt.to_excel(writer, sheet_name=m, startrow=6)
    writer.save()

    # Use openpyxl here to reference and change individual cells for titles
    wb = openpyxl.load_workbook(file)

    for sh in wb.worksheets:
        autosizexls(sh)

    for sh in wb.worksheets:
        sh['A1'].value = "Across All ROV"
        sh['A1'].font = Font(b=True, size=20)
        if sh.title not in ['Campaigns','Rooms']:
            sh['A2'].value = "Campaign State: " + sh.title
        sh['A2'].font = Font(b=True, size=16)

        if reportBy.lower() == "m":
            sh['A3'].value = "By Month"
        else:
            sh['A3'].value = "By Week"
        sh['A3'].font = Font(b=True, size=12)

        sh['A4'].value = "Date range, inclusive: " + minDate2 + " to " + maxDate2
        sh['A4'].font = Font(size=12)
        sh['A5'].value = "Source: " + VLdatafile
        sh['A5'].font = Font(size=12)

    wb.save(file)
    # Done with standalone enterprise wide report

    # Run reports at the room/org level
    orgs = xl['org_name'].unique()
    orgs[::-1].sort()

    for org in orgs:
        print("Org: " + org)
        xlo = xl[xl['org_name'] == org] # select records for one Org

        # Produce a workbook for each org with a filename like: "CA-North Coast VoterLetters Summary 2021-08-02.xlsx"
        file = os.path.join(os.path.expanduser(outputDirWFile),org + " VoterLetters Summary " + filedate + "-" + reportBy + ".xlsx")
        # Use Pandas dataframes here because it supports pivot tables and openpyxl does not

        teams = xlo['team_name'].unique()
        teams.sort()
        writer = pd.ExcelWriter(file)

        # Run standalone pivot on campaigns
        # df_pt = pd.pivot_table(xlo, columns=['data_week_of'], index=['parent_campaign_name', 'team_name'], values=['addresses_count'], margins=True, dropna=True,aggfunc=np.sum)
        df_pt = pd.pivot_table(xlo, columns=[reportVar], index=['master_campaign', 'parent_campaign_name'], values=['addresses_count'], margins=True, dropna=True,aggfunc=np.sum)
        df_pt.to_excel(writer, sheet_name='Campaigns', startrow=6) # summary sheet of al teams combined

        # Run standalone pivot on writers/teams
        df_pt = pd.pivot_table(xlo, columns=[reportVar], index=['team_name'], values=['addresses_count'], margins=True, dropna=True,
                               aggfunc=np.sum)
        df_pt.to_excel(writer, sheet_name='Team Sums', startrow=6) # summary sheet of al teams combined

        # df_pt = pd.pivot_table(xlo, columns=['team2'], index=['team_name', 'writer_name'], values=['addresses_count'], margins=True, dropna=True,
        df_pt = pd.pivot_table(xlo, columns=[reportVar], index=['team_name', 'writer_name'], values=['addresses_count'], margins=True, dropna=True,
                               aggfunc=np.sum)
        df_pt.to_excel(writer, sheet_name='Teams w Writers', startrow=6) # summary sheet of al teams combined

        # show count of requests to identify potential organizers (someone requesting small amounts and likely forwarding)
        # TODO: change var label on address_count to request_count
        df_pt = pd.pivot_table(xlo, columns=[reportVar], index=['team_name', 'writer_name'], values=['addresses_count'], margins=True, dropna=True,
                               aggfunc='count')
        df_pt.to_excel(writer, sheet_name='Teams w Counts', startrow=6) # summary sheet of al teams combined

        for t in teams:
            xlt = xlo[xlo['team_name'] == t]
            df_pt = pd.pivot_table(xlt, columns=[reportVar], index=['team_name', 'writer_name'], values=['addresses_count'], margins=True,
                                   dropna=True, aggfunc=np.sum)
            df_pt.to_excel(writer, sheet_name=t[:30], startrow=6) # had to take 30 chars of t because xl limit on sheet name length
        writer.save()

        # Use openpyxl here to reference and change individual cells for titles
        wb = openpyxl.load_workbook(file)

        # Insert a notes tab as first one. Must be openpyxl because excelwriter is dataframe only
        wb.create_sheet('Notes', 0)
        ws = wb["Notes"]
        r = 8
        for line in noteLines:
            ws.cell(row=r, column=1).value = line
            r = r + 1


        for sh in wb.worksheets:
            autosizexls(sh)
            # wb.active = 0 # did not work - still first two sheets activated!

        for sh in wb.worksheets:
         #   room = os.path.pathsplitext(InputFile)[0].split("/")[-1].split("_")[0]
            # print(os.path.split(file))
            sh['A1'].value = "Room: " + org
            sh['A1'].font = Font(b=True, size=20)
            # if sh.title != 'Campaigns':
            if sh.title not in ['Notes', 'Campaigns','Team Sums','Teams w Writers','Teams w Counts']:
                sh['A2'].value = "Team: " + sh.title
            sh['A2'].font = Font(b=True, size=16)

            if sh.title != 'All Teams':
                if reportBy.lower() == "m":
                    sh['A3'].value = "By Month"
                else:
                    sh['A3'].value = "By Week"
                sh['A3'].font = Font(b=True, size=12)

            sh['A4'].value = "Date range, inclusive: " + minDate2 + " to " + maxDate2
            sh['A4'].font = Font(size=12)
            sh['A5'].value = "Source: " + VLdatafile
            sh['A5'].font = Font(size=12)

            sh.sheet_view.tabSelected = False  # Deselects all via loop; defaults to index = 0 selected (active doesn't deselect any, only  activates!).

        wb.save(file)

        # wb = openpyxl.load_workbook(file)
        # wb.active = 0
        # wb.save(file)

    print("\nDone with all orgs.")
    pymsgbox.alert("Org reports produced. In:\n\n" + outputDir + "\n\n\nAdmin reports produced. In:\n\n" + outputDirAdmin, "Done!")

exit()
