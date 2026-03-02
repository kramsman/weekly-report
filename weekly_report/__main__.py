""" produce weekly reports on Sincere and upload them to Google drive.
Assign permissions to organizers which sends a Google notification.
 """

# FIXME: Import from the google_scripts directory set up the CFCG directory project.

# TODO: close 'select user list' window before running kicks off
# TODO: prompt for room report creation when Running?  Or does it not take that long, better to have all?
# TODO: pymsgbox prompt before tk openfile/dir because titles not showing
# TODO speed up the room report upload
# TODO: possible - add chart to all org reports (like admin)
# TODO: Move off my personal google account
# TODO: Add a tab or separate report that shows writers who have never requested.   How to avoid those invite to receive only and not order?
# TODO: add mapping email from organizer to requested email
# TODO: add excluded email list to skip giving permission
# TODO: format all numbers with commas
# TODO: remove pymsgbox references
# TODO Implement loguru with log file, levels of logging
# TODO when permission can't be granted on google sheet (recipient blocked) log msg and error to log file
# TODO: writer errors assigning permissions to a file rather than pymsgboxes during run

from weekly_report.create_report_files import create_report_files
from weekly_report.google_scripts.google_scripts import create_google_services
from weekly_report.google_scripts.google_scripts import upload_files

# FIXME below choked -
if False:  # DOES NOT WORK AS PACKAGE this updates the uvbekutils package which contains the little helper programs
    subprocess.run(["uv", "add", "uvbekutils", "--upgrade-package", "uvbekutils"], check=True)
    # uv add uvbekutils --upgrade-package uvbekutils
    #
    # from terminal
    # uv add uvbekutils@git+https://github.com/kramsman/uvbekutils.git
    # subprocess.run(["uv", "add","uvbekutils@git+https://github.com/kramsman/uvbekutils.git@optiona_showbuttons"],
    #                check=True)
    # subprocess.run(["uv", "add", "uvbekutils", "--upgrade-package", "uvbekutils@optiona_showbuttons"], check=True)

import datetime as dt
import inspect
# from apiclient.discovery import build
import logging
from loguru import logger

import numpy as np
import pandas as pd
# from apiclient.http import MediaFileUpload  # needed even if Pycharm says not
# import google.auth.transport.requests
from uvbekutils import setup_loguru, exit_yes
from uvbekutils import pyautobek
from constants import (
    ROOT_PATH, )

setup_loguru("DEBUG", "DEBUG")


logger.debug(f"({ROOT_PATH=}")


def check_sheet_headers(ws, vals):
    """
    Check list of (cell, val) tuples representing header labels in ws_to_chk and error if val not found in cell.
    eg vals = [('A1', 'use'), ('B1', 'fromFilePath'), ('C1', 'fromfilename'), ....]
    """

    def chk_header_vals(ws_to_chk, cell, val):
        """ error if val not found in wks cell. """
        if str(ws_to_chk[cell].value).strip().lower() != str(val).lower():
            exit_yes((f"Column heading '{cell}' on Setup sheet '{ws_to_chk.title}' not equal to literal '{val}'."
                      f"\n\nIt is '{str(ws_to_chk[cell].value)}'."),
                     )

    for pairs in vals:
        chk_header_vals(ws, pairs[0], pairs[1])


def check_df_headers(df, vals):
    """ Check list of values and compare to df column names.  Retun False if not matched. """

    vals = [val.lower().strip() for val in vals]
    cols = [col.lower().strip() for col in df.columns]

    if vals != cols:
        exit_yes((f"Dataframe columns do not match checked values.\n\n"
                  f"Columns are:\n\n"
                  f"'{','.join(cols)}\n\n"
                  f"Checked vals are:\n\n"
                  f"'{','.join(vals)}\n\n"
                  ),
                 )


def calling_func(level=0):
    """ returns the various levels of calling function.  0 is current, 1 is caller of current, etc """
    try:
        func = f"'{inspect.stack()[level][3]}', line #: {inspect.stack()[level][2]}"
    except Exception:
        func = f"** error ** inspect level too deep: {str(level)} called from {inspect.stack()[level][3]}"
    return func


def read_sincere_request_file(input_file):
    """ read the downloaded Sincere csv of requests and create df with all necessary variables"""
    request_df = pd.read_csv(input_file, na_filter=False)

    # request_df = request_df[request_df['org_name'].str.contains('Haar')]
    # request_df.to_excel('request_df.xlsx', index=False)

    # Check heading fields in Sincere request file to ensure fields didn't move/change
    check_df_headers(request_df,
                     ['parent_campaign_id', 'parent_campaign_name', 'request_id', 'created_at', 'writer_name',
                      'writer_email', 'addresses_count', 'org_name', 'org_id', 'team_name', 'factory_name', 'factory_id'])

    # replace '/' in some input fields (like org name) which cause issues when used as directories
    request_df['org_name'] = request_df['org_name'].str.replace("/", "-").str.replace("'", "-").str.replace(",", "-")
    request_df['team_name'] = request_df['team_name'].str.replace("/", "-").str.replace("'", "-").str.replace(",", "-")

    # todo: Sort orgs somewhere.  Here?

    # create a date field as a datetime object
    request_df['date2'] = pd.to_datetime(request_df['created_at'])

    # 'daysoffset' will contain the weekday 'day of week' as integers so we can step back to Monday
    request_df['daysoffset'] = request_df['date2'].apply(lambda x: x.weekday())
    # TODO Explain this timedelta operation
    request_df['data_week_ending'] = request_df.apply(lambda x: x['date2'] - dt.timedelta(days=x['daysoffset'] - 6), axis=1).dt.date
    request_df['month'] = pd.to_datetime(request_df['date2']).dt.to_period('M')

    # request_df.fillna(" No Team", inplace = True) # note space in front for sorting # change to only replace one col
    # ttt = request_df["team_name"].isnull() = " No Team"  # Fixme: THIS NEEDS FIXING
    request_df['team_name'] = np.where(request_df['team_name'].isnull() == True, " No Team", request_df['team_name'])
    request_df['team_name'] = request_df['team_name'].str.replace(":","-").str.replace("\\","-").str.replace("/","-")

    request_df = request_df[~request_df['org_name'].isin(['ROV Test Silo', 'ROV Training Silo', 'ROV Sample Silo'])]  # select records for one Org
    teams = sorted(request_df['team_name'].unique())


    request_df['master_campaign'] = np.where(request_df['factory_name'].isnull() == False, request_df['factory_name'],
                                     request_df['parent_campaign_name'].apply(lambda x: x[0:2]))

    # if factory name exists, trim parent campaign name  to just state/county
    request_df['parent_campaign_name'] = np.where(request_df['factory_name'].isnull() == True, request_df['parent_campaign_name'],
                                          request_df['parent_campaign_name'].str.split("-").str[0] + '-' +
                                          request_df['parent_campaign_name'].str.split("-").str[1])
    return request_df


def make_pivot_old(writer, df, report_var, index_vars, aggfunc, sheet_name, freeze_row, freeze_col):
    """ template for pivot tables in reports """
    df_pt = pd.pivot_table(df, columns=report_var, index=index_vars,
                           values=['addresses_count'], margins=True, dropna=True, aggfunc=aggfunc)
    df_pt = df_pt.sort_index(axis=1, ascending=False)
    df_pt.to_excel(writer, sheet_name=sheet_name, startrow=6)
    ws = writer.sheets[sheet_name]
    ws.freeze_panes(freeze_row, freeze_col)


def df_to_sheet(writer, df, sheet_name, freeze_cell=None):
    """ write df info to excel writer and freeze if specified """
    if sheet_name == '':
        sheet_name = "No Team"
    logger.debug(f"{sheet_name=}")
    df.to_excel(writer, sheet_name=sheet_name, startrow=6)
    ws = writer.sheets[sheet_name]
    # ws.freeze_panes(freeze_row, freeze_col)
    if freeze_cell is not None:
        mycell = ws[freeze_cell]
        ws.freeze_panes = mycell


def make_pivot(*, writer, df, report_var, index_vars, aggfunc, sheet_name, freeze_cell):
    """ template for pivot tables in reports

    Args:
        * ():
    """
    df_pt = pd.pivot_table(df, columns=report_var, index=index_vars,
                           values=['addresses_count'], margins=True, dropna=True, aggfunc=aggfunc)
    df_pt = df_pt.sort_index(axis=1, ascending=False)
    if sheet_name == '':
        sheet_name = "No Team"
    logger.debug(f"{sheet_name=}")
    df_pt.to_excel(writer, sheet_name=sheet_name, startrow=6)
    ws = writer.sheets[sheet_name]
    # ws.freeze_panes(freeze_row, freeze_col)
    mycell = ws[freeze_cell]
    ws.freeze_panes = mycell


def max_used_col(ws, row):
    """ Returns the column number (1 indexed) maximum non-none column in the input row of a sheet. """
    mxcol = 0
    for cell in reversed(ws[row]):
        if cell.value is not None:
            mxcol = cell.col_idx
            break
    return mxcol


def max_used_row(ws, col):
    """ Returns the column number (1 indexed) maximum non-none column in the input row of a sheet. """
    # for max_row, row in enumerate(ws, 1):
    #     if all(c.value is None for c in row):
    #         break
    mxrow = 0
    for cell in reversed(ws[col]):
        if cell.value is not None:
            mxrow = cell.row_idx
            break
    return mxrow


def make_chart(writer, df, index_vars, sheet_name):
    """ pivot sincere data by day/master then chart it """

    from openpyxl.chart import (
        LineChart,
        Reference,
    )

    # df_pt = pd.pivot_table(df, columns='created_at', index=index_vars,
    #                        values=['addresses_count'], margins=True, dropna=True, aggfunc=np.sum)
    df_pt = pd.pivot_table(df, columns='created_at', index=index_vars,
                           values=['addresses_count'], margins=True, dropna=True, aggfunc='sum')
    df_pt = df_pt.sort_index(axis=1, ascending=False)
    df_pt = df_pt.drop(df_pt.columns[0], axis=1)  # all column from margins=True (need all row)
    df_pt.to_excel(writer, sheet_name=sheet_name, startrow=6)
    ws = writer.sheets[sheet_name]
    # mycell = ws['C12']
    # ws.freeze_panes = mycell
    # ws.freeze_panes(freeze_row, freeze_col)

    # for r in dataframe_to_rows(df_pt, index=True, header=True):
    #     ws.append(r)

    c1 = LineChart()
    c1.title = "Daily Sincere Writer Address Requests"
    c1.style = 13
    c1.y_axis.title = 'Addresses'
    c1.x_axis.title = 'Date'

    # chart.y_axis.crossAx = 500
    # chart.x_axis = DateAxis(crossAx=100)
    # chart.x_axis.number_format ='yyyy/mm/dd'
    # chart.x_axis.majorTimeUnit = "days"

    used_col = max_used_col(ws,8)

    data = Reference(ws, min_col=1, max_col=used_col, min_row=10, max_row=ws.max_row)
    cats = Reference(ws, min_col=2, max_col=used_col, min_row=8, max_row=8)
    c1.add_data(data, from_rows=True, titles_from_data=True)
    c1.x_axis.number_format = 'mm/dd'

    c1.height = 10 * 2  # default is 7.5
    c1.width = 17 * 2  # default is 15
    c1.legend.position = 'b'  # 'tr', 'b', 't', 'l', 'r'

    c1.set_categories(cats)
    ws.add_chart(c1, "A1")
    # ws.add_chart(c1, "A1")
    # wb.save("SampleChart.xlsx")
    # writer.save()
    logger.debug("leaving make_chart")


def initialize():
    """  set up logger and if needed erase files, etc  """
    logging.root.handlers = []
    logging.getLogger().setLevel(100)
    logger_a = logging.getLogger('my_logger')
    logger_a_handler = logging.StreamHandler()
    logger_a_handler.setLevel(logging.DEBUG)
    logger_a_formatter = logging.Formatter('%(levelname)s - %(message)s')
    # logger_a_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s', "%Y-%m-%d %H:%M:%S")
    logger_a_handler.setFormatter(logger_a_formatter)
    logger_a.addHandler(logger_a_handler)

    return logger_a

def reminder():
    # Identify which VoterLetters files should be downloaded before starting
    if True:
        choice = pyautobek.confirm(
                      (f"\nDownload data from Sincere before running.\n\n"
                       f"1. Get addresses assigned in Sincere:\n\n"
                       f"   All Enterprise >\n"
                       f"   ROV >\n"
                       f"   Reports >\n"
                       f"   New Report >\n"
                       f"   Parent Campaign Address COUNTS\n\n"
                       f"   If 'Locked' is selected, make sure reports are not including erroneous campaigns.\n\n\n"
                       f"2. Get Master and County campaign information from Sincere:\n\n"
                       f"   Reports >\n"
                       f"   New Report >\n"
                       f"   All Parent Campaign Address REQUESTS\n\n"
                       f"   Dates 1/1/24 to prior Monday INCLUSIVE (includes to the day prior to specified).\n\n\n"
                       f"3. Get Sincere users (to assign google read permissions):\n\n"
                       f"   Reports >\n"
                       f"   New Report >\n"
                       f"   All USERS"),
                        "REMINDER",
                       buttons=["Ok", "Exit"],
                          )
        if choice == "exit":
            exit()

def main_program():
    """ create Sincere reports and upload them """

    my_logger = initialize()  # mostly set basic logger
    my_logger.setLevel(logging.DEBUG)  # level of logger to see on console
    logging.getLogger(name='my_logger').info(f"In {inspect.stack()[0][3]} - starting")

    drive_service, sheet_service = create_google_services()
    logging.getLogger(name='my_logger').debug(f"In {inspect.stack()[0][3]} - finished creating google services")

    # Identify which VoterLetters files should be downloaded before starting
    reminder()

    if True:
        choice = pyautobek.confirm(
                          (f"\n'Run' to create the admin report and room reports locally\n\n"
                               f"'Upload' to copy either the admin report, room reports, or both to Google sheets and send "
                               f"notifications to Organizers\n\n"
                               f"'Exit' to start over"
                               ),
                          'Run, Upload or Exit?',
                          buttons=["Run", 'Upload', 'Exit'])
    else:
        choice = 'run'

    if choice == "exit":
        exit()

    elif choice == "run":
        create_report_files()

    elif choice == "upload":
        upload_files(drive_service)


    logging.getLogger(name='my_logger').info(f"{inspect.stack()[0][3]} - All done!")
    exit()


if __name__ == '__main__':

    main_program()
