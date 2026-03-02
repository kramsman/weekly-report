"""Shared Excel and DataFrame utility helpers used across the weekly_report package."""

import inspect

import pandas as pd
from loguru import logger
from openpyxl.worksheet.worksheet import Worksheet
from uvbekutils import exit_yes
from uvbekutils import exit_yes


def check_sheet_headers(ws: Worksheet, vals: list[tuple[str, str]]) -> None:
    """Validate header cell values in an openpyxl worksheet against expected labels.

    Calls exit_yes() for each (cell, expected_value) pair where the cell's
    content does not match.

    Args:
        ws: The openpyxl worksheet to check.
        vals: List of (cell_address, expected_label) pairs,
            e.g. [('A1', 'use'), ('B1', 'fromFilePath')].
    """

    def chk_header_vals(ws_to_chk, cell, val):
        """Exit with an error message if a worksheet cell's value does not match an expected label.

        Args:
            ws_to_chk: The openpyxl worksheet to inspect.
            cell: Cell address string, e.g. 'A1'.
            val: Expected cell value.
        """
        if str(ws_to_chk[cell].value).strip().lower() != str(val).lower():
            exit_yes((f"Column heading '{cell}' on Setup sheet '{ws_to_chk.title}' not equal to literal '{val}'."
                      f"\n\nIt is '{str(ws_to_chk[cell].value)}'."),
                     )

    for pairs in vals:
        chk_header_vals(ws, pairs[0], pairs[1])


def check_df_headers(df: pd.DataFrame, vals: list[str]) -> None:
    """Validate that a DataFrame's column names exactly match an expected list.

    Comparison is case- and whitespace-insensitive. Calls exit_yes() if the
    columns do not match.

    Args:
        df: The DataFrame whose columns are checked.
        vals: Expected column names in the required order.
    """

    vals = [val.lower().strip() for val in vals]
    cols = [col.lower().strip() for col in df.columns]

    if vals != cols:
        exit_yes((f"Dataframe columns do not match checked values.\n\n"
                  f"Columns are:\n\n"
                  f"'{','.join(cols)}\n\n"
                  f"Checked vals are:\n\n"
                  f"'{','.join(vals)}\n\n"
                  ),
                 )


def calling_func(level: int = 0) -> str:
    """Return the name and source line of a function in the current call stack.

    Args:
        level: Stack depth to inspect. 0 is this function, 1 is its caller,
            2 is the caller's caller, etc. Defaults to 0.

    Returns:
        String of the form "'function_name', line #: <number>", or an error
        message if the requested stack level does not exist.
    """
    try:
        func = f"'{inspect.stack()[level][3]}', line #: {inspect.stack()[level][2]}"
    except Exception:
        func = f"** error ** inspect level too deep: {str(level)} called from {inspect.stack()[level][3]}"
    return func


def df_to_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, freeze_cell: str | None = None) -> None:
    """Write a DataFrame to an Excel sheet starting at row 7, with optional pane freeze.

    Args:
        writer: The active Excel writer.
        df: Data to write; output begins at row 7 (startrow=6).
        sheet_name: Target sheet name; substituted with 'No Team' if empty.
        freeze_cell: Cell reference at which to freeze panes, e.g. 'C10'.
            If None, panes are not frozen. Defaults to None.
    """
    if sheet_name == '':
        sheet_name = "No Team"
    logger.debug(f"{sheet_name=}")
    df.to_excel(writer, sheet_name=sheet_name, startrow=6)
    ws = writer.sheets[sheet_name]
    # ws.freeze_panes(freeze_row, freeze_col)
    if freeze_cell is not None:
        mycell = ws[freeze_cell]
        ws.freeze_panes = mycell


def max_used_col(ws: Worksheet, row: int) -> int:
    """Return the 1-based index of the last non-empty cell in a worksheet row.

    Args:
        ws: The openpyxl worksheet to inspect.
        row: 1-based row number to scan.

    Returns:
        1-based column index of the rightmost non-empty cell, or 0 if the
        row is entirely empty.
    """
    mxcol = 0
    for cell in reversed(ws[row]):
        if cell.value is not None:
            mxcol = cell.col_idx
            break
    return mxcol


def max_used_row(ws: Worksheet, col: str) -> int:
    """Return the 1-based index of the last non-empty cell in a worksheet column.

    Args:
        ws: The openpyxl worksheet to inspect.
        col: Column letter to scan, e.g. 'A'.

    Returns:
        1-based row index of the bottommost non-empty cell, or 0 if the
        column is entirely empty.
    """
    # for max_row, row in enumerate(ws, 1):
    #     if all(c.value is None for c in row):
    #         break
    mxrow = 0
    for cell in reversed(ws[col]):
        if cell.value is not None:
            mxrow = cell.row_idx
            break
    return mxrow
