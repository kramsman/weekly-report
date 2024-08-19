""" produce weekly reports on Sincere and upload them to Google drive.
Assign permissions to organizers which sends a Google notification.
Taken from 'Weekly VL Rpt V4.3.py' but no more version designations because using Git.
 """


# V3.0 includes Factory_name field/master campaign
#   4/26/22 if factory name exists, trim campaign to just state/county
# V3.2 6/16/22 skips over 3.1 which has Meika's requested reports and instead includes sending permission emails to the core groups
#  Also moves ROVwide report selection to the top
# V3.3 6-18-22 Go back to V3.1 and add in Meika's report: campaign by team/writer without date
# V3.4 Added "By Master" report to ROV overview
# V3.5 Fix error uploading file: trap error and report
#   convert files being uploaded from list to dataframe to make sorting and formatting easier
# V4.0 convert to subroutines
# V4.1 convert google API calls to routines
#   added chart
# V4.2 split admin reports and room to functions
# V4.3 replace pymsgbox with pysimplegui - include titles (which were lost in python v9+) and scrolling on boxes
#   Use upload prompts to only prompt for admin file and userfile & directory
#   Fixed assigning permissions to alternate gmails (with '+') by removing portion after +

# Google cloud console for authorizations: https://console.cloud.google.com/home/dashboard?project=voterletters-reports,
# then 'APi and Services' (upper left)>credentials.  +Create > Oauth > desktop
# delete token.json / credentials.json.  Rename to credentials, copy in

# TODO: close 'select user list' window before running kicks off
# TODO: prompt for room report creation when Running?  Or does it not take that long, better to have all?
# TODO: run another admin report- addresses in masters, avail, in rooms, with writers, avail
# TODO: pymsgbox prompt before tk openfile/dir because titles not showing
# TODO: include an "extra message" at the top of program that is appended to msg in permission email
# TODO speed up the room report upload
# TODO: possible - add chart to all org reports (like admin)
# TODO: Move off my personal google account
# TODO: Add a tab or separate report that shows writers who have never requested.   How to avoid those invite to receive only and not order?
# TODO: Can we make TK go to front so it doesn't get missed / crash?
# TODO: add mapping email from organizer to requested email
# TODO: add excluded email list to skip giving permission
# TODO: format all numbers with commas
# TODO: remove pymsgbox references
# TODO Implement loguru with log file, levels of logging
# TODO when permission can't be granted on google sheet (recipient blocked) log msg and error to log file


ROV_M_PERMISSION_MSG = (f"A new MONTHLY ROV-WIDE Sincere summary report has been sent to the CORE GROUP. "
                  f"Click to open.")
# ROV_W_PERMISSION_MSG = (f"A new WEEKLY ROV-WIDE Sincere summary report has been sent to the CORE GROUP. "
#                   f"Click to open.")
ROV_W_PERMISSION_MSG = ("A new WEEKLY ROV-WIDE Sincere summary report has been sent to the CORE GROUP. "
                        "\n\n** Note that weekly addresses to writers has jumped over the last 4 weeks "
                        "from 96K to 120K to 210K to 267K!!  Writers are taking over 20K per day!"
                        "\n** We are now out of FL addresses waiting on PDI.\n\n"
                        "Click to open.")

ORG_M_PERMISSION_MSG = ("A new MONTHLY Sincere summary report is available for your room. "
                  "To access the sheet you will need to be logged in to Google.  Do this by using the Chrome browser or by going to google.com in another browser."
                  "Click to open.")
ORG_W_PERMISSION_MSG = (f"A new WEEKLY Sincere summary report is available for your room. "
                  f"To access the sheet you will need to be logged in to Google.  Do this by using the Chrome browser or by going to google.com in another browser."
                  f"Click to open.")
ORG_W_PERMISSION_MSG = ("A new WEEKLY Sincere summary report is available for your room. "
                  "To access the sheet you will need to be logged in to Google.  Do this by using the Chrome browser or by going to google.com in another browser."
                        "\n\nWeekly addresses to writers has jumped 2 1/2 times over the last 4 weeks - "
                        "keep up the good work!!!"
                        "\n\n We are currently out of FL addresses waiting on our data vendor to supply them.\n\n"
                        "Click to open.")

# from datetime import datetime
import datetime as dt
import glob
import inspect
# from apiclient.discovery import build
import logging
import os
from loguru import logger
from pathlib import Path
# from pathlib import *
# from tkinter import *
from tkinter import Tk  # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename
import PySimpleGUI as sg
import sys

# from openpyxl import Workbook
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# from openpyxl.utils import column_index_from_string
# from random import random
import numpy as np
import openpyxl
import pandas as pd
# import xlsxwriter
import pymsgbox
# from apiclient.http import MediaFileUpload  # needed even if Pycharm says not
from googleapiclient.http import MediaFileUpload
# from tkinter import messagebox
# import ast
# import google.auth.transport.requests
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl.styles import Font
import shutil
import traceback
import re
from bekutils import setup_loguru, autosize_xls_cols, bad_path_create, exit_yes_no, exit_yes, \
    text_box, get_file_name, get_dir_name, check_ws_headers
from factory_and_campaign_subtotals import factory_and_campaign_subtotals

# log_level = "DEBUG"  # used for log file; screen set to INFO. TRACE, DEBUG, INFO, WARNING, ERROR
setup_loguru("DEBUG", "DEBUG")

# determine if application is running as a script file or frozen exe
if getattr(sys, 'frozen', False):
    # assume .exe is moved from subdirectory ./dist so reports go to same
    # ROOT_PATH = os.path.dirname(sys.executable)
    ROOT_PATH = Path(sys.executable).parents[0]
elif __file__:
    # ROOT_PATH = os.path.dirname(__file__)
    ROOT_PATH = Path(__file__).parents[0]
else:
    ROOT_PATH = None
# ROOT_PATH = str(ROOT_PATH)
logger.debug(f"({ROOT_PATH=}")

# string to filter factories.  Can not filter for 'not locked' because that would exclude some current year closed
# campaigns
FACTORY_FILTER_STRING = '-2024'

# for google drive api
SEND_PERMISSION_EMAIL_FLAG = True  # send permission granted emails
SCOPES = ['https://www.googleapis.com/auth/drive']
# ROOM_REPORT_FOLDER_ID = "1OEAdTnhQoAKpyzFqYdM1ztr42vHAqbk5"  # Folder holding Sincere Organizer Reports
ROOM_REPORT_FOLDER_ID = "1BcvXzwyEKGiiWSt27NLaassLE0DMgioP"  # 'VoterLetters Organizer Reports' copied to Tech@CFCG
# ADMIN_REPORT_FOLDER_ID = '1_vTXsMYOwK_TLYaRqC7YA_jolBuqDWP4'
# ADMIN_REPORT_FOLDER_ID = '1uILHx5lllvFqYjpzzYI-9JtTivHRitdx'  # test move folder from CFTech to Tech@CFCG
# ADMIN_REPORT_FOLDER_ID = '0AOkQpyahstHzUk9PVA'  # FAILED - would not give permission - share drive on Tech@CFCG
# ADMIN_REPORT_FOLDER_ID = '1hyArqMhrgCJeRoN7tHR8Rmkzn4BvCmpq'  # WORKS - Tech@CFCG my drive folder (NOT share)
ADMIN_REPORT_FOLDER_ID = '1PU8hcYfE3Vlh5v8Cq60Gup_8lfaFff4b'  # 'VoterLetters Admin Reports' copied to Tech@CFCG
SINCERE_DOWNLOAD_DIR = "~/Downloads/"
OUTPUT_DIR_ADMIN = "~/Dropbox/Postcard Files/VL Admin Reports"
OUTPUT_DIR_REPORTS = "~/Dropbox/Postcard Files/VL Org Reports"

CORE_EMAIL_LIST = ['kramsman@yahoo.com',
                   'Andrea@centerforcommonground.org',
                   'dee@centerforcommonground.org',
                   'comstockrov@gmail.com',
                   'nancy@centerforcommonground.org',
                   'bill.becky.rov@gmail.com',
                   'carey@harmonicsystems.net',
                   'gideon.asher1@gmail.com',
                   ]
# CORE_EMAIL_LIST = ['kramsman@yahoo.com', 'gideon.asher1@gmail.com']
# CORE_EMAIL_LIST = ['kramsman@yahoo.com']
# CORE_EMAIL_LIST = ['gideon.asher1@gmail.com']
# CORE_EMAIL_LIST = ['test@test.com', 'bkramer@kramericore.com']
# CORE_EMAIL_LIST = ['kramsman+test@Gmail.com']


noteLines = [
    '',
    'WHAT IS IN THIS REPORT',
    '',
    'What you’ll get:',
    '   - Each spreadsheet presents the address and request counts in your room: by time-period,'
    'state, campaign, team, and writer.',
    '   - The tabs (located at the bottom of the sheet) provide an overview then gets more specific:',
    '       the first few tabs show data for your whole room.  Tabs after that are one for each team.',
    '   - A weekly version showing counts by week will be produced each Monday. Monthly time periods',
    '       will be shown in a report produced on the first day of each month.',
    '   - You will still get the daily summary reports from Sincere.',
    '',
    'The overall report:',
    '   - The sheet shows only the address and request counts for your room.',
    '   - Each tab shows totals over time, either in weekly or monthly increments.  Dates are endings.',
    '   - The titles give the',
    '         - name of your room',
    '         - name of the team',
    '         - time increment (weekly or monthly)',
    '         - dates included in the report',
    '   - The data will be summarized in each row',
    '   - The reports show total state, campaign, teams, and writers',
    '',
    'The individual tabs of the report are:',
    '   - Tab 1, named  “Notes” explains the report'
    '   - Tab 2, named  “Campaigns” shows requested addresses broken by Campaign',
    '   - Tab 3, “All Team Sums” by Teams',
    '   - Tab 4, “Teams w Writers” by Writers within Teams',
    '   - Tab 5, “Teams w Counts” number of requests by Writers within Teams. ',
    '   - The tabs following are specific to each team (including individual writers).',
    '       These can be copied and pasted for team leads.  '
]


# def autosize_xls_cols(ws):
#     """ size all columns in worksheet to widest """
#     # from https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
#     dims = {}
#     for row in ws.rows:
#         for cell in row:
#             if cell.value:
#                 # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
#                 if cell.data_type == 'd':
#                     date_width = 10
#                 else:
#                     date_width = len(str(cell.value))
#                 dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), date_width))
#
#     for col, value in dims.items():
#         ws.column_dimensions[col].width = value + 1


# def get_creds_new():
#     """ get credentials needed to set up API service.  Was inline code; this cleaned things up."""
#     creds = None
#     # The file token.json stores the user's access and refresh tokens, and is
#     # created automatically when the authorization flow completes for the first
#     # time.
#     if not os.path.exists('token.json'):
#         creds = Credentials.from_authorized_user_file('token.json', SCOPES)
#         flow = InstalledAppFlow.from_client_secrets_file(
#             'credentials.json', SCOPES)
#         creds = flow.run_local_server(port=0)

#     # If there are no (valid) credentials available, let the user log in.
#     if not creds or not creds.valid:
#         if creds and creds.expired and creds.refresh_token:
#             os.remove('token.json')
#             creds.refresh(Request())
#         # else:
#             flow = InstalledAppFlow.from_client_secrets_file(
#                 'credentials.json', SCOPES)
#             creds = flow.run_local_server(port=0)

#         # Save the credentials for the next run
#         with open('token.json', 'w') as token:
#             token.write(creds.to_json())
#     return creds


# def get_creds(scopes):
    # """ get credentials needed to set up API service.  Was inline code; this cleaned things up."""
    # creds = None
    # # TODO this needs to be reworked if creds expired then delete token.json and force Google verification
    # # The file token.json stores the user's access and refresh tokens, and is
    # # created automatically when the authorization flow completes for the first
    # # time.
    # if os.path.exists('token.json'):
    #     creds = Credentials.from_authorized_user_file('token.json', scopes)
    # # If there are no (valid) credentials available, let the user log in.
    # if not creds or not creds.valid:
    #     if creds and creds.expired and creds.refresh_token:
    #         # os.remove('token.json')  # BEK 3/6/23 tried to fix by adding this but needs more
    #         creds.refresh(Request())
    #     else:
    #         flow = InstalledAppFlow.from_client_secrets_file(
    #             'credentials.json', SCOPES)
    #         creds = flow.run_local_server(port=0)

    #     # Save the credentials for the next run
    #     with open('token.json', 'w') as token:
    #         token.write(creds.to_json())
    # return creds
    

def get_creds(scopes, cred_file=None, cred_dir=None, token_file=None, token_dir=None, always_create=False,
              write_token=True):
    """ Gets credentials used in setting up google API services.  From webhook 1/6/24.

    Parameters
    ----------
    cred_dir :
    token_file :
    """

    import os
    # import pathlib
    # import pymsgbox
    # from google.auth.transport.requests import Request
    import google.auth.transport.requests
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    import pathlib

    def bek_cred_flow():
        logger.debug("get_creds using credentials -going to call 'InstalledAppFlow.from_client_secrets_file'")
        flow = InstalledAppFlow.from_client_secrets_file(cred_file, scopes)
        logger.debug(f"got flow: {flow.__dict__=}")
        # creds = flow.run_local_server(port=0)
        msg = ("Calling the 2nd part of flow, 'flow.run_local_server(port=0)'"
                 "\n\n   - Use 'TECH@CENTERFORCOMMONGROUND.ORG' google login."
                 "\n   - Must say process completed - close this window."
                 "\n   - If you get error 403, hit back in the browser and try again")
        logger.debug(msg)
        pymsgbox.alert(msg,"Verify Google Account")
        creds = flow.run_local_server(port=0)
        logger.debug(f"{creds.__dict__=}")
        return creds
            
    logger.debug(f"in get_creds with: {scopes=}, {cred_file=}, {cred_dir=}, {token_file=}, {token_dir=},"
                 f" {always_create=}, {write_token=}")
    if cred_dir is None:
        cred_dir = ROOT_PATH
    if cred_file is None:
        cred_file = 'credentials.json'
    cred_w_path = cred_dir / cred_file

    if token_dir is None:
        token_dir = ROOT_PATH
    if token_file is None:
        token_file = 'token.json'
    token_w_path = token_dir / token_file

    # token_w_path = Path(token_dir) / 'token.json'

    # SCOPES = ['https://www.googleapis.com/auth/drive']
    # TODO: Add code to delete token.json if creds fails.

    # FIXME this should check token (w refresh?) then credential
    if not cred_w_path.is_file():
        pymsgbox.alert(f"'{cred_w_path}' is missing. Copy it from another dir or download from API manager")
        # exit()

    if not token_w_path.is_file() and not cred_w_path.is_file():
        logger.error(f"'{token_w_path}' and {cred_w_path} are both missing.")
        pymsgbox.alert(f"'{token_w_path}' and {cred_w_path} are both missing.")
        exit()

    creds = None

    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first time.
    if token_w_path.is_file():
        # Credentials.from_authorized_user_file. Creates a Credentials instance from parsed authorized user info.
        # does it work from credentials, token, or either?
        # from_authorized_user_file - Creates a Credentials instance from an authorized user json file.
        creds = Credentials.from_authorized_user_file(token_w_path, scopes)
        logger.error(f"get_creds using token: {creds.__dict__=}")
        # logger.error(f"{dir(creds)=}")  # only names, not values, and includes builtins
        # logger.error(f"{vars(creds)=}")  # same as __dict__
        # logger.error(f"{help(creds)=}")
        # exit()

    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                # refresh(self, request) Refreshes the access token.
                # creds.refresh(Request())
                request = google.auth.transport.requests.Request()
                creds.refresh(request)
                logger.error("Creds refresh worked using token")
                logger.debug(f"{creds.__dict__=}")
            except Exception as e:
                pymsgbox.alert(f"Creds(request)) did not work using token.\nGoing to prompt for Google login."
                               f"\n\nError=\n{e}",
                               "Google Credential Issue")
                logger.debug("refresh failed - going to run bek_cred_flow")
                creds = bek_cred_flow()
        else:
            creds = bek_cred_flow()
            
        # Save the credentials for the next run
        if write_token:
            logger.debug("writing token")
            with open(token_w_path, 'w') as token:
                token.write(creds.to_json())
    logger.debug(f"ready to leave get_creds: {creds.__dict__=}")
    return creds


def get_sheet_values(sheet_service, sheet_range, header_strings):
    """ returns list of lists from a Google sheet range after checking the header strings are expected.
     Note: google sheet id is hardcoded via variable 'CFCG_LOOKUP_GOOGLE_SHEET_ID'."""
    logging.getLogger(name='my_logger').info(f"Starting {inspect.stack()[0][3]}")

    CFCG_LOOKUP_GOOGLE_SHEET_ID = 9999

    if isinstance(header_strings, str):
        header_strings_list = [field.strip().lower() for field in header_strings.split(",")]
    elif isinstance(header_strings, list):
        header_strings_list = [field.strip().lower() for field in header_strings]
    else:
        exit_yes(f"Headers passed to get_sheet_values is not string or list of headers.  It's "
                 f"{type(header_strings)}.")

    sheet_result = sheet_service.spreadsheets().values().get(spreadsheetId=CFCG_LOOKUP_GOOGLE_SHEET_ID, range=sheet_range).execute()
    sheet_values = sheet_result.get('values', [])
    headings = sheet_values.pop(0)  # read header and compare to ensure file format stays the same
    headings = [field.strip().lower() for field in headings]
    if headings != header_strings_list:
        exit_yes(f"Field names in {sheet_range} sheet header record are not {header_strings_list}; \n\nThey are {headings}")

    return sheet_values


def check_sheet_headers(ws, vals):
    """
    Check list of (cell, val) tuples representing header labels in ws_to_chk and error if val not found in cell.
    eg vals = [('A1', 'use'), ('B1', 'fromFilePath'), ('C1', 'fromfilename'), ....]
    """

    def chk_header_vals(ws_to_chk, cell, val):
        """ error if val not found in wks cell. """
        if str(ws_to_chk[cell].value).strip().lower() != str(val).lower():
            exit_yes((f"Column heading '{cell}' on Setup sheet '{ws_to_chk.title}' not equal to literal '{val}'."
                      f"\n\nIt is '{str(ws_to_chk[cell].value)}'."),
                     )

    for pairs in vals:
        chk_header_vals(ws, pairs[0], pairs[1])


def check_df_headers(df, vals):
    """ Check list of values and compare to df column names.  Retun False if not matched. """

    vals = [val.lower().strip() for val in vals]
    cols = [col.lower().strip() for col in df.columns]

    if vals != cols:
        exit_yes((f"Dataframe columns do not match checked values.\n\n"
                  f"Columns are:\n\n"
                  f"'{','.join(cols)}\n\n"
                  f"Checked vals are:\n\n"
                  f"'{','.join(vals)}\n\n"
                  ),
                 )


# def bad_path_create(path, msg=None):
#     """ checks for directory existence and creates if not found"""
#     if msg is None:
#         msg = ("Directory:\n\n" + path + "\n\ndoes not exist.  Creating." +
#                "\n\nCalled from " + calling_func(level=2))
#     if not os.path.isdir(path):
#         # pymsgbox.alert(msg, "Adding Directory via bad_path_create")
#         text_box("Alert", "", "\n\n" + msg, ["OK"])
#         os.makedirs(path)


def calling_func(level=0):
    """ returns the various levels of calling function.  0 is current, 1 is caller of current, etc """
    try:
        func = f"'{inspect.stack()[level][3]}', line #: {inspect.stack()[level][2]}"
    except Exception:
        func = f"** error ** inspect level too deep: {str(level)} called from {inspect.stack()[level][3]}"
    return func


# def exit_yes_no(msg, title=None, display_exiting=False):
#     """ makes this choice to continue one line"""
#     if not title:
#         title = "Exit?"
#     # choice = pymsgbox.confirm(msg, title, ['Yes', 'No'])
#     choice = text_box(title, "", "\n\n" + msg, ['Yes', 'No'])
#
#     if choice == "no":
#         if display_exiting:
#             # pymsgbox.alert("Exiting", "Alert")
#             text_box("Alert", "", "\n\nExiting", ["Ok"])
#         exit()


# def exit_yes(msg: str, title: str = None, *, errmsg: str = None) -> None:
#     """ exits program after giving user a popup window and raising an error. """
#     msg = (msg + "\n\n\nExiting." +
#            f"\n\nCalled from {calling_func(level=3)}"
#            f"\nCalled from {calling_func(level=2)}"
#            f"\nCalled from {calling_func(level=1)}"
#            )
#     if not errmsg:
#         errmsg = msg.replace("\n", " ")  # dont fill the console with linefeeds
#     if not title:
#         title = "** Exiting Program **"
#     # pymsgbox.alert(msg, title)
#     text_box(title, "", "\n\n" + msg, ["Ok"])
#     raise Exception(errmsg)



# def text_box(box_title, title2, txt, buttons=None):
#     """" Display text block with lines separated by \n and choice of buttons at bottom.
#     :param box_title: main heading on box
#     :type box_title: str
#     :param title2: 2nd title above text
#     :type title2: str
#     :param txt: text block with lines separated by \n
#     :type txt: str
#     :param buttons: list of button text, defaults to ['OK', 'Exit']
#     :type buttons: list of str
#     :return: lower case value of selected button
#     :rtype: str
#     """
#
#     if buttons is None:
#         buttons = ["OK", "Exit"]
#
#     col_factor = 3 # to scale windo equally
#     row_factor = 25 # to scale windo equally
#     max_cols = len(max(txt.split("\n"), key=len)) * col_factor
#     cols = max_cols
#     # v_scroll = False
#     col_limit = 80 * col_factor
#     col_min = 50 * col_factor
#     if cols > col_limit:
#         # v_scroll = True
#         cols = col_limit
#     elif cols < col_min:
#         cols = col_min
#
#
#     h_scroll = False
#     row_limit = 80
#     row_min = 10
#     max_rows = len(txt.split("\n"))
#     rows = max_rows
#     if rows > row_limit:
#         h_scroll = True
#         rows = row_limit
#     elif rows < row_min:
#         rows = row_min
#
#     layout = [
#         [sg.Text(title2,font=("Arial", 18))],
#         [sg.Multiline(txt, autoscroll=False, horizontal_scroll = h_scroll, expand_x=True,
#                       expand_y=True, enable_events=True )],
#         [sg.Button(text) for text in buttons],
#     ]
#
#     event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
#                               use_custom_titlebar=True, size=(600,rows*row_factor), disable_close=True,
#                               resizable=True, grab_anywhere	= True).read(close=True)
#     # event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
#     #                           use_custom_titlebar=True, size=(cols,rows), disable_close=True,
#     #                           resizable=True, grab_anywhere	= True).read(close=True)
#     if event is not None:
#         event = event.lower()
#     return event
#

# def get_dir_name(box_title, title2, initial_dir):
#     """ show an "Open" dialog box and return the selected directory. Replaced askdirectory with pyeasygui
#     :param title2:
#     :type title2:
#     """
#
#     layout = [
#         [sg.Text(title2,font=("Arial", 18))],
#         [
#          sg.Input(key="-IN-", expand_x=True),
#          sg.FolderBrowse(initial_folder=os.path.expanduser(initial_dir))
#          ],
#         [sg.Button("Choose")],
#     ]
#
#     # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
#     event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
#                               size=(1000,150), use_custom_titlebar=True ).read(close=True)
#
#     dir_name = values['-IN-']
#     if dir_name == "":
#         exit_yes("No directory name chosen")
#
#     return dir_name
#
#
# def get_file_name(box_title, title2, initial_dir):
#     """ show an "Open" dialog box and return the selected file name. Replaced askopenfilename with pyeasygui
#     :param title2: heading of the box
#     :type title2: text next to input field
#     """
#     #"Select Sincere address export file 'all-parent-campaign-requests-yyyy-mm-dd.csv'"
#     layout = [
#         [sg.Text(title2,font=("Arial", 18))],
#         [
#          sg.Input(key="-IN-", expand_x=True),
#          sg.FileBrowse(initial_folder=os.path.expanduser(initial_dir))
#          ],
#         [sg.Button("Choose")],
#     ]
#
#     # event, values = sg.Window(heading_in_box, layout, size=(600, 100)).read(close=True)
#     event, values = sg.Window(box_title, layout, titlebar_font=("Arial", 20), font=("Arial", 14),
#                               size=(1000, 150), use_custom_titlebar=True).read(close=True)
#     # sg.Window.close()
#
#     file_name = values['-IN-']
#     if file_name == "":
#         exit_yes("No file name chosen")
#
#     return file_name
#

def read_sincere_request_file(input_file):
    """ read the downloaded Sincere csv of requests and create df with all necessary variables"""
    request_df = pd.read_csv(input_file, na_filter=False)

    # Check heading fields in Sincere request file to ensure fields didn't move/change
    check_df_headers(request_df,
                     ['parent_campaign_id', 'parent_campaign_name', 'request_id', 'created_at', 'writer_name',
                      'writer_email', 'addresses_count', 'org_name', 'org_id', 'team_name', 'factory_name', 'factory_id'])

    # replace '/' in some input fields (like org name) which cause issues when used as directories
    request_df['org_name'] = request_df['org_name'].str.replace("/", "-").str.replace("'", "-").str.replace(",", "-")
    request_df['team_name'] = request_df['team_name'].str.replace("/", "-").str.replace("'", "-").str.replace(",", "-")

    # todo: Sort orgs somewhere.  Here?

    # create a date field as a datetime object
    request_df['date2'] = pd.to_datetime(request_df['created_at'])

    # 'daysoffset' will contain the weekday 'day of week' as integers so we can step back to Monday
    request_df['daysoffset'] = request_df['date2'].apply(lambda x: x.weekday())
    # TODO Explain this timedelta operation
    request_df['data_week_ending'] = request_df.apply(lambda x: x['date2'] - dt.timedelta(days=x['daysoffset'] - 6), axis=1).dt.date
    request_df['month'] = pd.to_datetime(request_df['date2']).dt.to_period('M')

    # request_df.fillna(" No Team", inplace = True) # note space in front for sorting # change to only replace one col
    # ttt = request_df["team_name"].isnull() = " No Team"  # Fixme: THIS NEEDS FIXING
    request_df['team_name'] = np.where(request_df['team_name'].isnull() == True, " No Team", request_df['team_name'])
    request_df['team_name'] = request_df['team_name'].str.replace(":","-").str.replace("\\","-").str.replace("/","-")

    request_df = request_df[~request_df['org_name'].isin(['ROV Test Silo', 'ROV Training Silo', 'ROV Sample Silo'])]  # select records for one Org
    teams = request_df['team_name'].unique()
    teams.sort()

    state_list = ["AL", "AK", "AS", "AZ", "AR", "CA", "CO", "CT", "DE", "DC", "FL", "GA", "GU", "HI", "ID", "IL", "IN",
                  "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM",
                  "NY", "NC", "ND", "MP", "OH", "OK", "OR", "PA", "PR", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA",
                  "VI", "WA", "WV", "WI", "WY"]

    request_df['master_campaign'] = np.where(request_df['factory_name'].isnull() == False, request_df['factory_name'],
                                     request_df['parent_campaign_name'].apply(lambda x: x[0:2]))

    # if factory name exists, trim parent campaign name  to just state/county
    request_df['parent_campaign_name'] = np.where(request_df['factory_name'].isnull() == True, request_df['parent_campaign_name'],
                                          request_df['parent_campaign_name'].str.split("-").str[0] + '-' +
                                          request_df['parent_campaign_name'].str.split("-").str[1])
    return request_df


def make_pivot_old(writer, df, report_var, index_vars, aggfunc, sheet_name, freeze_row, freeze_col):
    """ template for pivot tables in reports """
    df_pt = pd.pivot_table(df, columns=report_var, index=index_vars,
                           values=['addresses_count'], margins=True, dropna=True, aggfunc=aggfunc)
    df_pt = df_pt.sort_index(axis=1, ascending=False)
    df_pt.to_excel(writer, sheet_name=sheet_name, startrow=6)
    ws = writer.sheets[sheet_name]
    ws.freeze_panes(freeze_row, freeze_col)


def df_to_sheet(writer, df, sheet_name, freeze_cell=None):
    """ write df info to excel writer and freeze if specified """
    if sheet_name == '':
        sheet_name = "No Team"
    logger.debug(f"{sheet_name=}")
    df.to_excel(writer, sheet_name=sheet_name, startrow=6)
    ws = writer.sheets[sheet_name]
    # ws.freeze_panes(freeze_row, freeze_col)
    if freeze_cell is not None:
        mycell = ws[freeze_cell]
        ws.freeze_panes = mycell


def make_pivot(writer, df, report_var, index_vars, aggfunc, sheet_name, freeze_cell):
    """ template for pivot tables in reports """
    df_pt = pd.pivot_table(df, columns=report_var, index=index_vars,
                           values=['addresses_count'], margins=True, dropna=True, aggfunc=aggfunc)
    df_pt = df_pt.sort_index(axis=1, ascending=False)
    if sheet_name == '':
        sheet_name = "No Team"
    logger.debug(f"{sheet_name=}")
    df_pt.to_excel(writer, sheet_name=sheet_name, startrow=6)
    ws = writer.sheets[sheet_name]
    # ws.freeze_panes(freeze_row, freeze_col)
    mycell = ws[freeze_cell]
    ws.freeze_panes = mycell


def max_used_col(ws, row):
    """ Returns the column number (1 indexed) maximum non-none column in the input row of a sheet. """
    mxcol = 0
    for cell in reversed(ws[row]):
        if cell.value is not None:
            mxcol = cell.col_idx
            break
    return mxcol


def max_used_row(ws, col):
    """ Returns the column number (1 indexed) maximum non-none column in the input row of a sheet. """
    # for max_row, row in enumerate(ws, 1):
    #     if all(c.value is None for c in row):
    #         break
    mxrow = 0
    for cell in reversed(ws[col]):
        if cell.value is not None:
            mxrow = cell.row_idx
            break
    return mxrow


def make_chart(writer, df, index_vars, sheet_name):
    """ pivot sincere data by day/master then chart it """
    from openpyxl import Workbook

    from openpyxl.chart import (
        LineChart,
        Reference,
    )

    # df_pt = pd.pivot_table(df, columns='created_at', index=index_vars,
    #                        values=['addresses_count'], margins=True, dropna=True, aggfunc=np.sum)
    df_pt = pd.pivot_table(df, columns='created_at', index=index_vars,
                           values=['addresses_count'], margins=True, dropna=True, aggfunc='sum')
    df_pt = df_pt.sort_index(axis=1, ascending=False)
    df_pt = df_pt.drop(df_pt.columns[0], axis=1)  # all column from margins=True (need all row)
    df_pt.to_excel(writer, sheet_name=sheet_name, startrow=6)
    ws = writer.sheets[sheet_name]
    # mycell = ws['C12']
    # ws.freeze_panes = mycell
    # ws.freeze_panes(freeze_row, freeze_col)

    # for r in dataframe_to_rows(df_pt, index=True, header=True):
    #     ws.append(r)

    c1 = LineChart()
    c1.title = "Daily Sincere Writer Address Requests"
    c1.style = 13
    c1.y_axis.title = 'Addresses'
    c1.x_axis.title = 'Date'

    # chart.y_axis.crossAx = 500
    # chart.x_axis = DateAxis(crossAx=100)
    # chart.x_axis.number_format ='yyyy/mm/dd'
    # chart.x_axis.majorTimeUnit = "days"

    used_col = max_used_col(ws,8)

    data = Reference(ws, min_col=1, max_col=used_col, min_row=10, max_row=ws.max_row)
    cats = Reference(ws, min_col=2, max_col=used_col, min_row=8, max_row=8)
    c1.add_data(data, from_rows=True, titles_from_data=True)
    c1.x_axis.number_format = 'mm/dd'

    c1.height = 8.5 * 2  # default is 7.5
    c1.width = 11 * 2  # default is 15
    c1.legend.position = 'b'  # 'tr', 'b', 't', 'l', 'r'

    c1.set_categories(cats)
    ws.add_chart(c1, "A1")
    # ws.add_chart(c1, "A1")
    # wb.save("SampleChart.xlsx")
    # writer.save()


def create_admin_report(sincere_df, sincere_data_file, report_by, str_output_dir_admin, admin_rpt_filename, factory_csv,
                        df_sort_func):
    """ create admin reports and chart.  Sincere data file just for report titles.

    Args:
        df_sort_func ():
        factory_csv ():
        sincere_df (): df containing df  address request info
        sincere_data_file (): name of file used to create sincere_df, used for report title.  like
        'all-parent-campaigns-requests-2024-02-26 test.csv'
        report_by (): W or M for Weekly or Monthly
        str_output_dir_admin (): str of directory where admin reports will be placed
        admin_rpt_filename (): name of the xls file for admin report
    """

    print("Org: Enterprise wide")  # TODO: use logger instead of prints

    report_by = report_by.upper()
    if report_by == 'M':
        report_var = "month"
    elif report_by == 'W':
        report_var = "data_week_ending"
    else:
        exit_yes(f"Report_by not W or M: {report_by}")

    output_dir_admin = Path(str_output_dir_admin).expanduser()

    excel_output_file = Path(output_dir_admin).expanduser() / admin_rpt_filename

    writer = pd.ExcelWriter(excel_output_file, engine="openpyxl")

    # create chart of daily requests
    logger.debug("calling make_chart")
    make_chart(writer, sincere_df, 'factory_name', 'Chart')

    # Report on Masters only using sum w subtotals
    factory_tots, factory_pull_date = factory_and_campaign_subtotals(factory_csv, FACTORY_FILTER_STRING,
                                                                     break_fields="[('Election', True), ('Factory', "
                                                                                  "False), ]")

    factory_tots.sort_index(level=['Election', 'Factory'], key=df_sort_func, ascending=False, inplace=True)
    df_to_sheet(writer, factory_tots, 'Assigned_by_state', freeze_cell="C8")

    # Report on Masters and Campaigns using sum w subtotals
    logger.debug("calling factory_and_campaign_subtotals")
    factory_tots, factory_pull_date = factory_and_campaign_subtotals(factory_csv, FACTORY_FILTER_STRING,
                                                                     break_fields="[('Election', True), ('Factory', "
                                                                                  "True), ('Name', False), ]")

    factory_tots2 = factory_tots.sort_index(level=['Election', 'Factory', 'Name'], key=df_sort_func, ascending=False)
    df_to_sheet(writer, factory_tots2, 'Assigned_w_counties', freeze_cell="D8")

    # Run pivot on Master without county campaigns
    logger.debug("calling make_pivot")
    make_pivot(writer, sincere_df, [report_var], ['master_campaign'], 'sum', 'Masters', 'B10')

    # Run standalone pivot on campaigns
    logger.debug("calling make_pivot")
    make_pivot(writer, sincere_df, [report_var], ['master_campaign', 'parent_campaign_name'], 'sum', 'Campaigns', 'C10')


    logger.debug("calling make_pivot")
    make_pivot(writer, sincere_df, [report_var], ['org_name'], 'sum', 'Rooms', 'B10')

    wb = writer.book

    # size columns before titles added because they are very wide
    for sh in wb.worksheets:
        autosize_xls_cols(sh)

    min_date2 = sincere_df['created_at'].min()
    max_date2 = sincere_df['created_at'].max()
    for sh in wb.worksheets:
        sh['A1'].value = "Across All ROV"
        sh['A1'].font = Font(b=True, size=20)
        if sh.title not in ['Campaigns', 'Rooms', 'Masters', 'Totals']:
            sh['A2'].value = "Campaign State: " + sh.title
        sh['A2'].font = Font(b=True, size=16)

        sh['A3'].value = ("By Month" if report_by == "M" else "By Week")
        sh['A3'].font = Font(b=True, size=12)

        if sh.title in ['Totals']:  # factory / campaign snapshot file
            sh['A4'].value = f"Data as of: {factory_pull_date}"
        else:
            sh['A4'].value = "Date range, inclusive: " + min_date2 + " to " + max_date2

        sh['A4'].font = Font(size=12)

        if sh.title in ['Totals']:  # factory / campaign snapshot file
            sh['A5'].value = "Source: " + str(factory_csv.name)
        else:
            sh['A5'].value = "Source: " + sincere_data_file
        sh['A5'].font = Font(size=12)

    # writer.save()
    writer.close()


def create_room_reports(sincere_df, sincere_data_file, file_date, report_by, str_output_dir_rooms ):
    """ create all the room reports """
    report_by = report_by.upper()
    if report_by == 'M':
        report_var = "month"
    elif report_by == 'W':
        report_var = "data_week_ending"
    else:
        exit_yes(f"Report_by not W or M: {report_by}")

    orgs = sincere_df['org_name'].unique()
    orgs[::-1].sort()

    for org in orgs:
        print("Org: " + org)
        xlo = sincere_df[sincere_df['org_name'] == org]  # select records for one Org

        # Name workbook for each org with like: "CA-North Coast VoterLetters Summary 2021-08-02.xlsx"
        room_file_name = f"{org} Sincere Summary {file_date}-{report_by}.xlsx"

        excel_output_file = os.path.join(os.path.expanduser(str_output_dir_rooms),
                            org + " Sincere Summary " + file_date + "-" + report_by + ".xlsx")
#    outputDirWFile = os.path.join(outputDir,Path(InputFile).stem + "-" + reportBy)


        teams = xlo['team_name'].unique()
        teams.sort()
        writer = pd.ExcelWriter(excel_output_file, engine='openpyxl')

        # pivot on campaigns
        # make_pivot(writer, xlo, [report_var], ['master_campaign', 'parent_campaign_name'], np.sum, 'Campaigns', 'C10')
        make_pivot(writer, xlo, [report_var], ['master_campaign', 'parent_campaign_name'], 'sum', 'Campaigns', 'C10')

        # pivot on writers/teams
        # make_pivot(writer, xlo, [report_var], ['team_name'], np.sum, 'Team Sums', 'B10')
        make_pivot(writer, xlo, [report_var], ['team_name'], 'sum', 'Team Sums', 'B10')

        # make_pivot(writer, xlo, [report_var], ['team_name', 'writer_name'], np.sum, 'Teams w Writers', 'C10')
        make_pivot(writer, xlo, [report_var], ['team_name', 'writer_name'], 'sum', 'Teams w Writers', 'C10')

        # show COUNT (rather than np.sum) of requests to identify potential organizers
        make_pivot(writer, xlo, [report_var], ['team_name', 'writer_name'], 'count', 'Teams w Counts', 'C10')
        # TODO: change var label on address_count to request_count

        # Meika's two reports - cols by campaign not date
        # make_pivot(writer, xlo, ['master_campaign', 'parent_campaign_name'], ['team_name'], np.sum, 'Team by Campaigns', 'B11')
        make_pivot(writer, xlo, ['master_campaign', 'parent_campaign_name'], ['team_name'], 'sum', 'Team by Campaigns', 'B11')

        # make_pivot(writer, xlo, ['master_campaign', 'parent_campaign_name'], ['team_name', 'writer_name'], np.sum,
        #            'Teams w Writers by Campaign', 'B11')
        make_pivot(writer, xlo, ['master_campaign', 'parent_campaign_name'], ['team_name', 'writer_name'], 'sum',
                   'Teams w Writers by Campaign', 'B11')

        # # TODO: sort campaigns by master latest date (which?) then campaign latest date, not alpha

        for team in teams:
            xlt = xlo[xlo['team_name'] == team]
            shname = team[:30]
            # make_pivot(writer, xlt, [report_var], ['team_name', 'writer_name'], np.sum, shname, 'C10')
            make_pivot(writer, xlt, [report_var], ['team_name', 'writer_name'], 'sum', shname, 'C10')

        # Insert a notes tab as first one. Must be openpyxl because excelwriter is dataframe only
        wb = writer.book
        wb.create_sheet('Notes', 0)
        ws = wb["Notes"]
        r = 8
        for line in noteLines:
            ws.cell(row=r, column=1).value = line
            r = r + 1

        for sh in wb.worksheets:
            autosize_xls_cols(sh)
            # wb.active = 0 # did not work - still first two sheets activated!

        min_date2 = sincere_df['created_at'].min()
        max_date2 = sincere_df['created_at'].max()
        for sh in wb.worksheets:
            sh['A1'].value = "Room: " + org
            sh['A1'].font = Font(b=True, size=20)
            if sh.title not in ['Notes', 'Campaigns', 'Team Sums', 'Teams w Writers', 'Teams w Counts']:
                sh['A2'].value = "Team: " + sh.title
            sh['A2'].font = Font(b=True, size=16)

            if sh.title != 'All Teams':
                sh['A3'].value = ("By Month" if report_by.lower() == "m" else "By Week")
                sh['A3'].font = Font(b=True, size=12)

            sh['A4'].value = "Date range, inclusive: " + min_date2 + " to " + max_date2
            sh['A4'].font = Font(size=12)
            sh['A5'].value = "Source: " + sincere_data_file
            sh['A5'].font = Font(size=12)

            sh.sheet_view.tabSelected = False  # Deselects all via loop; defaults to index = 0 selected (active doesn't deselect any, only  activates!).

        writer.close()
        # writer.save()


def create_report_files():
    """ produce all the spreadsheet reports locally"""

    report_by = text_box("\n\n[W]eekly or [M]onthly report?", "Date Format", "", ["W", "M", 'Exit'])
    report_by = report_by.upper()
    if report_by == 'M':
        report_var = "month"
    elif report_by == 'W':
        report_var = "data_week_ending"
    else:
        exit()

    input_file = get_file_name("Pick a File",
                               "Select Sincere address export file 'all-parent-campaign-REQUESTS-yyyy-mm-dd'",
                               SINCERE_DOWNLOAD_DIR)

    factory_csv = get_file_name("Pick File",
                                f"Pick a parent-campaign file to summarize (eg "
                                f"'parent-campaign-address-counts-2023-08-03.csv'."
                                f"\n\nCreate via ROV > Reports > New Report > Parent Campaign Address COUNTS",
                                SINCERE_DOWNLOAD_DIR)

    # Create dataframe of excel sheet data
    sincere_df = read_sincere_request_file(input_file)
    sincere_data_file_name = os.path.basename(input_file)

    # only take factories from this year - many old and some new are locked
    sincere_df = sincere_df.loc[sincere_df['factory_name']
        .apply(lambda x: (FACTORY_FILTER_STRING in x.lower()))]
    # remove any unwanted requests - training and sample rooms
    sincere_df = sincere_df.loc[sincere_df['factory_name']
        .apply(lambda x: not ('zzz' in x.lower() or 'xxx' in x.lower() or 'test' in x.lower()))]
    # remove more unwanted requests - training and sample rooms
    sincere_df = sincere_df.loc[sincere_df['org_name']
        .apply(lambda x: not ('zzz' in x.lower() or 'xxx' in x.lower() or 'training' in x.lower() or 'sample' in
                         x.lower()))]

    # Set Election field to one of two values: General if 'general' found in name else Primary
    sincere_df['election'] = sincere_df['factory_name'].apply(lambda fact: ('General' if 'general' in fact.lower()
                                                                            else 'Primary'))

    # Create a dictionary of the earliest(min) date in a Factory to use for sorting
    factory_dict = sincere_df.groupby(by=['factory_name'], dropna=False)['created_at'].min().apply(
        pd.to_datetime).dt.strftime('%Y%m%d').to_dict()
    factory_dict['_TOTAL'] = 999999

    file_date = Path(input_file).stem[-10:]

    master_campaigns = sincere_df['master_campaign'].unique()
    master_campaigns.sort()

    ## add dicts to 'level_dicts'; dict will be used for sort values for and dict name matching level name in multiindex
    ## working version to play can be found in sort df using dict.py
    level_dicts = dict()  # name of 'level_dicts' is hardcoded in sort function
    # Key value in level_dicts is matched to level name in index (case-insensitive)
    level_dicts['Factory'] = factory_dict
    level_dicts = {k.lower(): v for k, v in level_dicts.items()}

    # define multiindex_df_sorter function used to sort multiiindex output of groupby by dictionaries
    def multiindex_df_sorter(level, default_level_dict={'_TOTAL': 999999}):
        """ function for sorting a dataframe's multiindex using a dictionary for each level.  if name of index
        (case-insensitive) field matches key in dict hardcoded as 'level_dicts' then dictionary is used,
        otherwise index level is left as is via use of an empty dictionary.  """
        level_dict = level_dicts.get(level.name.lower(), default_level_dict)
        mapped_index = level.map(lambda index_item: level_dict.get(index_item, index_item))
        return mapped_index

    ### Create admin report
    admin_rpt_filename = f"ROVWide Sincere Summary {file_date}-{report_by}.xlsx"
    # create_admin_report(sincere_df, sincere_data_file_name, report_by, OUTPUT_DIR_ADMIN, admin_rpt_filename,
    #                     factory_csv)
    create_admin_report(sincere_df, sincere_data_file_name, report_by, OUTPUT_DIR_ADMIN, admin_rpt_filename,
                        factory_csv, multiindex_df_sorter)


    # Base of report file; input file name is used to create a sub dir under this
    output_dir = os.path.expanduser(OUTPUT_DIR_REPORTS)
    output_dir_w_file = os.path.join(output_dir, Path(input_file).stem + "-" + report_by)
    # CHECK and create dir if not exists
    # bad_path_create(output_dir_w_file)
    # if not os.path.isdir(output_dir_w_file):
    #     print(f"Adding Directory {output_dir_w_file}\n")
    #     os.makedirs(output_dir_w_file)
    if os.path.isdir(output_dir_w_file):
        shutil.rmtree(output_dir_w_file)
        print(f"Removing directory {output_dir_w_file}\n")
    os.makedirs(output_dir_w_file)
    # Run reports at the room/org level
    create_room_reports(sincere_df, sincere_data_file_name, file_date, report_by, output_dir_w_file)

    print("\nDone with all orgs.")
    text_box(f"Org reports produced. In:"
               f"\n\n{OUTPUT_DIR_REPORTS}"
               f"\n\n\nAdmin reports produced. In:"
               f"\n\n{OUTPUT_DIR_ADMIN}","Done!", "",
             ["Ok"])


def upload_sheet_to_drive(drive_service, file_to_upload_w_path, drive_folder_id):
    """ upload a local file to a google drive.  Returns fileid of created file """
    # allow one folder id or list to be passed
    if isinstance(drive_folder_id, str):
        drive_folder_id = [drive_folder_id]

    file_name_wo_ext = Path(file_to_upload_w_path).stem
    file_metadata = {'name': file_name_wo_ext,
                     'mimeType': 'application/vnd.google-apps.spreadsheet',
                     "parents": drive_folder_id}

    # converts xlsx to google sheet
    try:
        media = MediaFileUpload(file_to_upload_w_path,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except:
        print(f"**** ERROR CONVERTING XLS FILE TO GOOGLE SHEET: Uploaded file name:  {file_name_wo_ext}, "
              f"path: {file_to_upload_w_path}")
        # pymsgbox.alert(f"**** CONVERTING XLS FILE TO GOOGLE SHEET: Uploaded file name:  {file_to_upload_w_path}",
        #            "CHECK PYTHON CONSOLE FOR ERROR")
        text_box(f"\n\n**** CONVERTING XLS FILE TO GOOGLE SHEET: Uploaded file name: "
                     f" {file_to_upload_w_path}","CHECK PYTHON CONSOLE FOR ERROR", "",
                 ["Ok"])

    try:
        file = drive_service.files().create(body=file_metadata,
                                        supportsAllDrives=True,
                                        media_body=media,
                                        fields='id').execute()
    except:
        # if file upload doesn't work
        print(f"**** ERROR WITH UPLOAD: Uploaded file name: {file_name_wo_ext}, path: "
          f"{file_to_upload_w_path}")
        # pymsgbox.alert(f"**** ERROR WITH UPLOAD: Uploaded file name: {file_name_wo_ext}",
        #            "CHECK PYTHON CONSOLE FOR ERROR")
        text_box(f"\n\n**** ERROR WITH UPLOAD: Uploaded file name: {file_name_wo_ext}",
                 "CHECK PYTHON CONSOLE FOR ERROR", "",
                 ["Ok"])

    uploaded_file_id = file.get('id')

    return uploaded_file_id


def permission_to_drive_file(drive_service, drive_file_id, email_flag, email, permission_msg):
    """ grant user permission to a google drive file.  optionally notify user with google generated email and custom
    message"""
    # use regex to remove all text between '+' and '@gmail' cause fails in permission grant
    email_clean = re.sub('\+[^>]+@gmail.com', '@gmail.com', email, flags=re.IGNORECASE)
    if email_clean != email:
        print(f"'{email}' changed to '{email_clean}' for permission grant.")
        logging.getLogger(name='my_logger').info(f"In {inspect.stack()[0][3]} - '{email}' changed to '{email_clean}' "
                                                 f"for permission grant.")


    payload = {
        "role": "reader",
        "type": "user",
        "emailAddress": email_clean
    }

    if not email_flag:
        permission_msg = None

    # TODO: need to trap adding permission when notification off because people without google accounts error.
    try:
        # If email is not gmail and doesn't have an associated gmail account, notification must be true, below
        drive_service.permissions().create(fileId=drive_file_id,
                                           supportsAllDrives=True,
                                            sendNotificationEmail=email_flag,
                                            emailMessage=permission_msg,
                                            body=payload).execute()
    except Exception as e:
        print('First Except', sys.exc_info()[2])
        traceback.print_exc()
        # Try again with notification set to true
        try:
            drive_service.permissions().create(fileId=drive_file_id,
                                       sendNotificationEmail=True,
                                       emailMessage=permission_msg,
                                       body=payload).execute()
        except Exception as e:
            print(f"Error giving permission, click ok to continue: Email: {email}, Uploaded file id: {drive_file_id}")
            print(sys.exc_info()[2])
            traceback.print_exc()
            # pymsgbox.alert(
            #     f"*****  ERROR GIVING PERMISSION: Email: {email}, Uploaded file id: {drive_file_id}",
            #     "CHECK PYTHON CONSOLE FOR ERROR")
            text_box(f"Error giving permission, ok to ignore: (likely blocked by recipient): Email: "
                         f"{email}, Uploaded file id: {drive_file_id}",
                     "Click OK to Continue", "",
                     ["Ok"])

    a=1


def upload_admin_report(drive_service, admin_report_to_upload):
    """ upload admin report from local to google drive
    """
    # do the upload
    admin_report_to_upload = Path(admin_report_to_upload)
    file_name_wo_ext = Path(admin_report_to_upload).stem  # filename without extension

    # delete all files with same name in the destination admin folder
    # delete_file_list1 = get_google_file_ids(drive_service, file_name_wo_ext, ADMIN_REPORT_FOLDER_ID)
    delete_file_list = get_google_file_or_folder_ids(drive_service, 'file', file_name_wo_ext, ADMIN_REPORT_FOLDER_ID)
    delete_list_of_google_files(drive_service, delete_file_list)

    print('Uploading filename ', file_name_wo_ext)
    uploaded_file_id = upload_sheet_to_drive(drive_service, admin_report_to_upload, ADMIN_REPORT_FOLDER_ID)

    # set permissions and email admin report to Core group
    if not SEND_PERMISSION_EMAIL_FLAG:
        permission_msg = None
    elif file_name_wo_ext[-2:].lower() == '-m':
        permission_msg = ROV_M_PERMISSION_MSG
        # permission_msg = (f"A new MONTHLY ROV-WIDE Sincere summary report has been sent to the CORE GROUP. "
        #                   f"Click to open.")
    else:
        permission_msg = ROV_W_PERMISSION_MSG
        # permission_msg = (f"A new WEEKLY ROV-WIDE Sincere summary report has been sent to the CORE GROUP. "
        #                   f"Click to open.")

    for email in CORE_EMAIL_LIST:
        # print('  - Adding ROV-wide permission for email: ', email)
        logging.getLogger(name='my_logger').info(f"{inspect.stack()[0][3]}  - Adding ROV-wide permission for email: "
                                                 f"', {email}")
        permission_to_drive_file(drive_service, uploaded_file_id, SEND_PERMISSION_EMAIL_FLAG, email, permission_msg)


def get_google_file_or_folder_ids(drive_service, file_or_folder, folder_name, parent):
    """ create list of all folder or file ids in parent which match name """
    if file_or_folder.lower() == 'folder':
        file_or_folder_compare = '='
    elif file_or_folder.lower() == 'file':
        file_or_folder_compare = '!='
    else:
        exit_yes(f"Google drive file_or_folder parameter of {file_or_folder} not 'file' or folder'")

    query_string = (f"mimeType {file_or_folder_compare} 'application/vnd.google-apps.folder' and "
                    f"trashed=false and "
                    f"name='{folder_name}' and " 
                    f"parents = '{parent}'"
                    )
    # id_list = []
    page_token = None
    while True:
        response = drive_service.files().list(
            q=query_string,
            spaces='drive',
            fields="nextPageToken, files(id, name, trashed, parents)",
            pageToken=page_token).execute()
        print(f"Found old folders with same name- response: {response}")


        id_list = [file.get('id') for file in response.get('files', [])]

        # for file in response.get('files', []):
        #     print((f"Found folder- name: {file.get('name')}, id: {file.get('id')}, "
        #           f"trashed: {file.get('trashed')}, parents: {file.get('parents')}"))

        page_token = response.get('nextPageToken', None)
        if page_token is None:
            break
    return id_list


def delete_list_of_google_files(drive_service, file_id_list):
    """ delete all fileids in list from Google drive """
    for folderId in file_id_list:
        drive_service.files().update(fileId=folderId, body={'trashed': True}).execute()
        # trash instead of delete moves to trash; delete gone for good


def create_google_services():
    """ create and returns a Google service for drive and sheet so we can read from each """
    creds = get_creds(SCOPES, cred_file='credentials.json', cred_dir=ROOT_PATH, token_file='token.json', token_dir=ROOT_PATH, always_create=False,
              write_token=True)
    try:
        drive_service = build('drive', 'v3', credentials=creds)
        sheet_service = build('sheets', 'v4', credentials=creds)
    except HttpError as err:
        print(err)

    return drive_service, sheet_service


def create_drive_subfolder(drive_service, folder_name, drive_folder_id):
    """ create a subfolder on a google drive given the id of the parent """
    if isinstance(drive_folder_id, str):
        drive_folder_id = [drive_folder_id]

    file_metadata = {
        'name': folder_name,
        'parents': drive_folder_id,
        'mimeType': 'application/vnd.google-apps.folder'}
    file = drive_service.files().create(body=file_metadata, fields='id').execute()
    new_folder_id = file.get('id')
    # print('Folder ID created: %s' % file.get('id'))

    return new_folder_id


def upload_room_reports(drive_service, str_dir_to_upload, organizer_email_list):
    """ upload all room reports in directory and send notification via Google sheet permission
    """

    report_dir_to_upload = Path(str_dir_to_upload)
    folder_name = report_dir_to_upload.parts[-1]  # the last section of the path # TODO change to better function than parts

    # id list of folders to trash(delete)
    # delete_folder_list1 = get_google_folder_ids(drive_service, folder_name, ROOM_REPORT_FOLDER_ID)
    delete_folder_list = get_google_file_or_folder_ids(drive_service, 'folder', folder_name,
                                                             ROOM_REPORT_FOLDER_ID)
    # TODO maybe combine delete folder and file and delete_list_of_google_files
    delete_list_of_google_files(drive_service, delete_folder_list)

    # Create sub folder
    new_folder_id = create_drive_subfolder(drive_service, folder_name, ROOM_REPORT_FOLDER_ID)

    # get list of filenames in computer directory
    excel_report_files_lst = glob.glob(str(report_dir_to_upload / '[!~]*.xlsx'))
    # xlsx files not starting with ~ (temporary files)

    report_files_df = pd.DataFrame(excel_report_files_lst, columns=['file_w_path'])
    report_files_df['file_name'] = report_files_df['file_w_path'].map(lambda x: os.path.basename(x))
    report_files_df['file_name_wo_ext'] = report_files_df['file_w_path'].map(lambda x: Path(x).stem)
    report_files_df['room_name'] = report_files_df['file_name_wo_ext'].map(lambda x: x.split(' Sincere ')[0])
    # report_files_df['room_name'] = report_files_df['file_name_wo_ext'].map(lambda x: x[0:-34])

    report_files_df.sort_values(by=['room_name'], ascending=True, inplace=True)

    for index, row in report_files_df.iterrows():
        file_w_path, file_name, file_name_wo_ext, room_name = row
        # # assign report_files_df row to variables to avoid report_files_df(loc,'field') notation
        print(f"\nUploading filename:  {file_name}, room_name: {room_name}")

        uploaded_file_id = upload_sheet_to_drive(drive_service, file_w_path, new_folder_id)

        for org, email in organizer_email_list:
            # email = 'briank@kramericore.com'
            if org.replace(",", "-") == room_name:  # had to replace , with - in file names so this makes room match.
                print(f"  - Adding permission for room_name: {room_name}, email: {email}")

                # set permissions and email admin report to Core group
                if not SEND_PERMISSION_EMAIL_FLAG:
                    permission_msg = None
                elif file_name_wo_ext[-2:].lower() == '-m':
                    permission_msg = ORG_M_PERMISSION_MSG
                    # permission_msg = ("A new MONTHLY VoterLetters summary report is available for your room. "
                    #                   "To access the sheet you will need to be logged in to Google.  Do this by using the Chrome browser or by going to google.com in another browser."
                    #                   "Click to open.")
                else:
                    permission_msg = ORG_W_PERMISSION_MSG
                    # permission_msg = (f"A new WEEKLY VoterLetters summary report is available for your room. "
                    #                   f"To access the sheet you will need to be logged in to Google.  Do this by using the Chrome browser or by going to google.com in another browser."
                    #                   f"Click to open.")

                # print('  - Adding ROV-wide permission for email: ', email)
                # logging.getLogger(name='my_logger').info(f"{inspect.stack()[0][3]} - Adding room permission for email: "
                #     f"{email}")
                # pymsgbox.confirm("Wait a bit and then try")
                permission_to_drive_file(drive_service, uploaded_file_id, SEND_PERMISSION_EMAIL_FLAG, email, permission_msg)
                a=1


def upload_files(drive_service):
    """ upload report files from local drive to google drive and alert recipients via google permissions
    """

    # choice = pymsgbox.confirm("Upload admin report?", "Upload Admin Reports?", ['Yes', 'No'])
    choice = text_box("\n\nUpload Admin Reports?", "Upload admin report?", "",  ['Yes', 'No'])
    upload_admin = False
    if choice == "yes":
        upload_admin = True
        if True:  # False for testing w hardcoded admin file
            # pymsgbox.alert("Select ROV Admin report to upload to Core")
            str_admin_report_to_upload = get_file_name("Select ROV ADMIN report to upload to Core Group",
                                                       "",
                                                       OUTPUT_DIR_ADMIN)
        else:
            str_admin_report_to_upload = "/Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/VL Admin " \
                                          "Reports/TEST ROVWide Sincere Summary 2023-03-14-W.xlsx"


    # choice = pymsgbox.confirm("Upload room reports?", "Upload Room Reports?", ['Yes', 'No'])
    choice = text_box("\n\nUpload Room Reports?", "Upload room reports?", "", ['Yes', 'No'])
    upload_room = False
    if choice == "yes":
        upload_room = True
        if True:  # False for testing w report directory hardcoded file
            # pymsgbox.alert("Select report directory to UPLOAD")
            str_report_dir_to_upload = get_dir_name("Select a REPORT DIRECTORY to upload",
                                                    "",
                                                    OUTPUT_DIR_REPORTS)
        else:
            # str_report_dir_to_upload = "/Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/VL Org Reports/TEST all-parent-campaigns-requests-2023-03-14-W"
            str_report_dir_to_upload = "/Users/Denise/Library/CloudStorage/Dropbox/Postcard Files/VL Org Reports/all-parent-campaigns-requests-2023-05-15 GIDEON-W"

        if True:  # True prompts for users file to read; false uses hardcoded test file
            # pymsgbox.alert("Select Sincere export file for USER LIST (all-users-yyyy-mm-dd.csv)")
            sincere_user_file = get_file_name("Select Sincere export file for USER LIST (all-users-yyyy-mm-dd.csv)",
                                              "",
                                              SINCERE_DOWNLOAD_DIR)
            df = pd.read_csv(sincere_user_file, usecols=['name', 'email', 'role', 'is_active', 'organization'])
        else:
            df = pd.read_csv('/Users/Denise/Downloads/all-users-2023-05-15 GIDEON.csv',
                             usecols=['name', 'email', 'role', 'is_active', 'organization'])

        df = df.sort_values(['organization', 'name'], ascending=(True, True))
        organizer_email_list = df.loc[(df['role'].isin(['organizer']) & (df['is_active'])),
                                       ['organization', 'email']].values.tolist()
        # organizer_email_list = [
        #     ['NY-NYC and Long Island', 'kramsman+nycorg@gmail.com'], ['NY-NYC and Long Island', 'bkramer@kramericore.com']
        # ]


    if upload_admin:
        # do the upload from local to google sheets
        upload_admin_report(drive_service, str_admin_report_to_upload)

    if upload_room:
        # do the upload from local to google sheets
        upload_room_reports(drive_service, str_report_dir_to_upload, organizer_email_list)


def initialize():
    """  set up logger and if needed erase files, etc  """
    logging.root.handlers = []
    logging.getLogger().setLevel(100)
    logger_a = logging.getLogger('my_logger')
    logger_a_handler = logging.StreamHandler()
    logger_a_handler.setLevel(logging.DEBUG)
    logger_a_formatter = logging.Formatter('%(levelname)s - %(message)s')
    # logger_a_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s', "%Y-%m-%d %H:%M:%S")
    logger_a_handler.setFormatter(logger_a_formatter)
    logger_a.addHandler(logger_a_handler)

    return logger_a


def main_program():
    """ create Sincere reports and upload them """

    my_logger = initialize()  # mostly set basic logger
    my_logger.setLevel(logging.DEBUG)  # level of logger to see on console
    logging.getLogger(name='my_logger').info(f"In {inspect.stack()[0][3]} - starting")

    drive_service, sheet_service = create_google_services()
    logging.getLogger(name='my_logger').debug(f"In {inspect.stack()[0][3]} - finished creating google services")

    # Identify which VoterLetters files should be downloaded before starting
    if True:
        choice = text_box(
                      (f"\nDownload data from Sincere before running.\n\n"
                       f"1. Get addresses assigned in Sincere:\n\n"
                       f"   All Enterprise >\n"
                       f"   ROV >\n"
                       f"   Reports >\n"
                       f"   New Report >\n"
                       f"   Parent Campaign Address COUNTS\n\n"
                       f"   If 'Locked' is selected, make sure reports are not including erroneous campaigns.\n\n\n"
                       f"2. Get Master and County campaign information from Sincere:\n\n"
                       f"   Reports >\n"
                       f"   New Report >\n"
                       f"   All Parent Campaign Address REQUESTS\n\n"
                       f"   Dates 1/1/24 to prior Monday INCLUSIVE (includes to the day prior to specified).\n\n\n"
                       f"3. Get Sincere users (to assign google read permissions):\n\n"
                       f"   Reports >\n"
                       f"   New Report >\n"
                       f"   All USERS\n\n\n"),
                        "REMINDER", "",
                       buttons=["Ok", "Exit"],
                          )
        if choice == "exit":
            exit()

    choice = text_box(
                      (f"\n'Run' to create the admin report and room reports locally\n\n"
                           f"'Upload' to copy either the admin report, room reports, or both to Google sheets and send "
                           f"notifications to Organizers\n\n"
                           f"'Exit' to start over"
                           ),
                      'Run, Upload or Exit?', "",
                      buttons=["Run", 'Upload', 'Exit'])
    if choice == "exit":
        exit()

    elif choice == "run":
        create_report_files()

    elif choice == "upload":
        upload_files(drive_service)


    logging.getLogger(name='my_logger').info(f"{inspect.stack()[0][3]} - All done!")
    exit()


if __name__ == '__main__':

    main_program()
