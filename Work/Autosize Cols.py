# test import an outide file into a program
# routine autosize_xls_cols sets column widths on a workbook opened with openpyxl

import xlsxwriter
import datetime
from datetime import datetime
import pymsgbox
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from random import random
import numpy as np
import pandas as pd

from tkinter import *
from tkinter import Tk  # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory
from tkinter import messagebox
from uszipcode import SearchEngine
from math import floor
from pandasgui import show

# from UliPlot.XLSX import auto_adjust_xlsx_column_width

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


z = 1.1E3
t = type(z).__name__
x = t in ['complex', 'float', 'int']

def get_col_widths(dataframe):
    # First we find the maximum length of the index column
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    # return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]
    zz = [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]
    return zz

def best_fit_cols(dataframe,wks):
    for i, width in enumerate(get_col_widths(dataframe)):
        wks.column_dimensions[get_column_letter(i + 1)].width = width

# from https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
def autosizexls(ws):
    print(ws.title)
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                print("cell value", cell.value, "data_type",cell.data_type)
                # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

    for col, value in dims.items():
        print(col , value)
        ws.column_dimensions[col].width = value + 1

collist = ['A', 'B']

# IP = pd.DataFrame({'A':[13.1,"","a", 3345],'B':["x", "y", "z","q"],'C':["a", "b", "c","d"]})
IP = pd.DataFrame({'lastname':["a","","a", "b"],'address':["1111xdfgdfgdfgddfg8888", "y", "z","q"],'score':[1.2234, 654.666,45, 1.1]})
# IP = pd.DataFrame([1,"x"],[213.1,"y"],[3,"z"], columns=['A','B'])
df_pt = pd.pivot_table(IP, index=['address'], values=['score','lastname'], aggfunc='count', margins=True)

# IP['key']= (IP['lastname'].str[:30] + '#' + IP['address'] + '#' + IP['city']).str.lower().str.replace(' ','').str.replace('-','').str.replace('\'','')

file = os.path.expanduser(os.path.join("~/Dropbox/Postcard Files/TestInputFiles/Campaigns/TEST VA MERGE TRUNCATED/USPS CLEAN/Formatted", "Summary", "Test format.xlsx"))
# writer = pd.ExcelWriter(file)
writer = pd.ExcelWriter(file, engine='openpyxl')

# Code below taken from web but not tested  https://techoverflow.net/2021/03/05/how-to-auto-fit-pandas-pd-to_excel-xlsx-column-width/
# pip install UliPlot
# # Export dataset to XLSX
# with pd.ExcelWriter(file) as writer:
df_pt.to_excel(writer, sheet_name='AA', startrow=4)
df_pt.to_excel(writer, sheet_name='BB', startrow=5)
# df_pt.to_excel(writer, sheet_name="MySheet")
# auto_adjust_xlsx_column_width(df_pt, writer, sheet_name="RawData", margin=0)
writer.save()

# wb = xlsxwriter.Workbook(file)
# ws = wb.get_worksheet_by_name('AA')
# ws = wb.active

wb = openpyxl.load_workbook(file)
tdy = datetime.now()

autosizexls(wb['AA'])
wb.save(file)

# sheets = ['AA', 'BB']
sheets = ['AA']

for sheet in sheets:
    # for ws in wb.worksheets:
    # for ws in wb[sheet]:
    ws = wb[sheet]
    # ws = wb['AA']
    # x = ws['A5'].value
    # lx = len(ws['A5'].value)
    # y = ws['A6'].value
    # ly = len(ws['A6'].value)
    # best_fit_cols(df_pt, ws)
    ws["A1"] = "Test Report really  really  really  really  really  long"
    ws['A1'].font = Font(b=True, size=16)
    ws["A2"] = ws.title
    ws['A2'].font = Font(b=True, size=12)
    ws["A3"] = datetime.now().strftime('%m/%d/%Y')




# ws.column_dimensions[get_column_letter(0 + 1)].width = 22
# ws.column_dimensions["A"].width = 22

# ws.set_column(0,0, 22)
# ws.set_column(1,1, 8)
# ws.set_column(2,2, 5)

wb.save(file)
# writer.save()
# ws = wb['AA']


a=1
