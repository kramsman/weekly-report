# summarizes data by room and shows those with more small writers


from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from datetime import datetime
import datetime as dt
import xlsxwriter
import pymsgbox
import os
from pathlib import *
import pathlib
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from random import random
import numpy as np
import pandas as pd
from tkinter import *
from tkinter import Tk  # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory
from tkinter import messagebox
import ast
import google.auth.transport.requests
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from apiclient.http import MediaFileUpload

smallLimit = 60 # cutoff for small request
largelimit = 1000 # cutoff for large request

def autosizexls(ws):
    # from https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    # print(ws.title)
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                # print("cell value", cell.value, "data_type",cell.data_type)
                if(cell.data_type == 'd'):
                    dateWidth = 10
                else: dateWidth = len(str(cell.value))
                # dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), dateWidth))

    for col, value in dims.items():
        # print(col , value)
        ws.column_dimensions[col].width = value + 1


def smallWriterReport():
    #### Start Main code
    ####
    # Identify which VoterLetters files should be downloaded before starting
    # pymsgbox.alert("Download data from VoterLetters before running.\n\n"\
    #                "1. Addresses assigned:\n      VoterLetters All Enterprise > ROV >\n      Reports >\n      New Report >\n      All Parent Campaign Address Requests.\n  Dates 1/1/22 to prior Monday INCLUSIVE (includes to the day prior to specified).\n\n"\
    #                "2. Users (to assign google read permissions):\n      VoterLetters Child Organizations >\n      Export Users to CSV.\n",\
    #                "Get ready!")

    # pymsgbox.alert("** Running Here  **", "Alert")

    # outputDirAdmin = os.path.expanduser("~/Dropbox/Postcard Files/VL Admin Reports")

    # Dir to prompt to start looking for input file
    inputDir = os.path.expanduser("/Users/Denise/Downloads/")
    # Base of report file; input file name is used to create a sub dir under this
    # outputDir = os.path.expanduser("~/Dropbox/Postcard Files/VL Org Reports")

    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing

    # InputFile = askopenfilename(initialdir=os.path.expanduser(r"/Users/Denise/Downloads/"),title="Select VoterLetters address export file", filetypes=(("CSV files", "*.csv "),))  # show an "Open" dialog box and return the path to the selected file
    InputFile = askopenfilename(initialdir=os.path.expanduser(inputDir),title="Select VoterLetters address export file (all-parent-campaigns-requests-yyyy-mm-dd.csv)", filetypes=(("CSV files", "*.csv "),))  # show an "Open" dialog box and return the path to the selected file
    # tried but failed InputFile = askopenfilename(initialdir=os.path.expanduser(inputDir),title="Select VoterLetters address export file", filetypes=('VL files','all-parent*.*'))  # show an "Open" dialog box and return the path to the selected file
    # TODO: set filter to only display fileNAMES that are proper input, such as 'all-parent*.csv'

    # head, tail = os.path.split(InputFile)
    VLdatafile = os.path.basename(InputFile)
    # outputDirWFile = os.path.join(outputDir,Path(InputFile).stem + "-" + reportBy)
    # filedate = Path(InputFile).stem[-10:]

    # CHECK and create dir if not exists
    # if not os.path.isdir(outputDirWFile):
    #     # pymsgbox.alert("Input Path does not exist. Creating.\n\n" + outputDirWFile, "Fixing Directory Structure")
    #     # FIXME: close Tkinter window here
    #     os.makedirs(outputDirWFile)

    # Create dataframe of excel sheet data
    xl = pd.read_csv(InputFile)

    # # replace '/' in some input fields (like org name) which cause issues when used as directories
    xl['org_name'] = xl['org_name'].str.replace("/","-").str.replace("'","-").str.replace(",","-")
    xl = xl[~xl['org_name'].isin(['ROV Test Silo','ROV Training Silo','ROV Sample Silo'])]  # delete test orgs

    # xl['team_name'] = xl['team_name'].str.replace("/","-").str.replace("'","-").str.replace(",","-")

    # # create a date field as a datetime object
    # xl['date2'] = pd.to_datetime(xl['created_at'])
    minDate2 = xl['created_at'].min()
    maxDate2 = xl['created_at'].max()

    xl['small'] = np.where( xl['addresses_count'] <= smallLimit, 1, 0 )
    xl['small_quant'] = xl['addresses_count'] * xl['small']
    xl['large'] = np.where(xl['addresses_count'] >= largelimit, 1, 0)
    xl['large_quant'] = xl['addresses_count'] * xl['large']

    xl = xl.rename(columns={'request_id' : 'total_requests'})

    file = "ROVWide Room Summary Showing Small Writers.xlsx"

    # Use Pandas dataframes here because it supports pivot tables and openpyxl does not
    # now use writer because it works with pandas dataframes
    writer = pd.ExcelWriter(file)

    df_pt = pd.pivot_table(xl, index=['org_name'],
                           values=[ 'addresses_count','total_requests', 'small', 'large','small_quant','large_quant'],
                           aggfunc={'addresses_count': 'sum',
                                    'total_requests': 'count',
                                    'small': 'sum',
                                    'small_quant': 'sum',
                                    'large': 'sum',
                                    'large_quant': 'sum'
                                    },
                           margins=True, dropna=True )
    df_pt['pct_small'] = round((df_pt['small'] / df_pt['total_requests'] ) * 100,1)
    df_pt['pct_large'] = round((df_pt['large'] / df_pt['total_requests'] ) * 100,1)
    df_pt['pct_small_quant'] = round((df_pt['small_quant'] / df_pt['addresses_count'] ) * 100,1)
    df_pt['pct_large_quant'] = round((df_pt['large_quant'] / df_pt['addresses_count'] ) * 100,1)
    df_pt.sort_values(['addresses_count'], ascending=[False], inplace=True)
    # df.sort(['a', 'b'], ascending=[True, False])
    df_pt = df_pt[["addresses_count", "total_requests", "small", "large","pct_small", "pct_large",'small_quant','large_quant','pct_small_quant','pct_large_quant']]

    # aggfunc = {'D': np.mean,
    #            'E': [min, max, np.mean]}

    df_pt.to_excel(writer, sheet_name='Rooms', startrow=6)  # summary sheet of al teams combined
    writer.save()

    # Use openpyxl here to reference and change individual cells for titles
    wb = openpyxl.load_workbook(file)

    for sh in wb.worksheets:
        autosizexls(sh)

    for sh in wb.worksheets:
        sh['A1'].value = "ROV Address Request by Room Showing Request Size"
        sh['A1'].font = Font(b=True, size=20)
        sh['A3'].value = "Small are <= " + str(smallLimit) + ", Large >= " + str(largelimit)
        sh['A3'].font = Font(b=True, size=12)
        sh['A4'].value = "Date range, inclusive: " + minDate2 + " to " + maxDate2
        sh['A4'].font = Font(size=12)
        sh['A5'].value = "Source: " + VLdatafile
        sh['A5'].font = Font(size=12)

    wb.save(file)
    # Done with standalone enterprise wide report

    a=1

if __name__ == '__main__':
    smallWriterReport()

a=1
