# summarizes data by room and shows those with more small writers
# THIS WAS SCRAPPED - harder than worth to only use openpyxl, especially given index bug


from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import datetime as dt
import xlsxwriter
import pymsgbox
import os
from pathlib import *
import pathlib
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from random import random
import numpy as np
import pandas as pd
from tkinter import *
from tkinter import Tk  # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory
from tkinter import messagebox
import ast
import google.auth.transport.requests
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from apiclient.http import MediaFileUpload

smallLimit = 100 # cutoff for small request
medlimit = 250 # cutoff for medium request

def autosizexls(ws):
    # from https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    # print(ws.title)
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                # print("cell value", cell.value, "data_type",cell.data_type)
                if(cell.data_type == 'd'):
                    dateWidth = 10
                else: dateWidth = len(str(cell.value))
                # dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), dateWidth))

    for col, value in dims.items():
        # print(col , value)
        ws.column_dimensions[col].width = value + 1


def smallWriterReport():
    #### Start Main code
    ####
    # Identify which VoterLetters files should be downloaded before starting
    # pymsgbox.alert("Download data from VoterLetters before running.\n\n"\
    #                "1. Addresses assigned:\n      VoterLetters All Enterprise > ROV >\n      Reports >\n      New Report >\n      All Parent Campaign Address Requests.\n  Dates 1/1/22 to prior Monday INCLUSIVE (includes to the day prior to specified).\n\n"\
    #                "2. Users (to assign google read permissions):\n      VoterLetters Child Organizations >\n      Export Users to CSV.\n",\
    #                "Get ready!")

    # pymsgbox.alert("** Running Here  **", "Alert")

    # outputDirAdmin = os.path.expanduser("~/Dropbox/Postcard Files/VL Admin Reports")

    # Dir to prompt to start looking for input file
    inputDir = os.path.expanduser("/Users/Denise/Downloads/")
    # Base of report file; input file name is used to create a sub dir under this
    # outputDir = os.path.expanduser("~/Dropbox/Postcard Files/VL Org Reports")

    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing

    # InputFile = askopenfilename(initialdir=os.path.expanduser(r"/Users/Denise/Downloads/"),title="Select VoterLetters address export file", filetypes=(("CSV files", "*.csv "),))  # show an "Open" dialog box and return the path to the selected file
    if False:
        InputFile = askopenfilename(initialdir=os.path.expanduser(inputDir),title="Select VoterLetters address export file (all-parent-campaigns-requests-yyyy-mm-dd.csv)", filetypes=(("CSV files", "*.csv "),))  # show an "Open" dialog box and return the path to the selected file
    InputFile = '/Users/Denise/Downloads/all-parent-campaigns-requests-2022-05-09.csv'
    # tried but failed InputFile = askopenfilename(initialdir=os.path.expanduser(inputDir),title="Select VoterLetters address export file", filetypes=('VL files','all-parent*.*'))  # show an "Open" dialog box and return the path to the selected file
    # TODO: set filter to only display fileNAMES that are proper input, such as 'all-parent*.csv'

    # head, tail = os.path.split(InputFile)
    VLdatafile = os.path.basename(InputFile)
    # outputDirWFile = os.path.join(outputDir,Path(InputFile).stem + "-" + reportBy)
    # filedate = Path(InputFile).stem[-10:]

    # CHECK and create dir if not exists
    # if not os.path.isdir(outputDirWFile):
    #     # pymsgbox.alert("Input Path does not exist. Creating.\n\n" + outputDirWFile, "Fixing Directory Structure")
    #     # FIXME: close Tkinter window here
    #     os.makedirs(outputDirWFile)

    # Create dataframe of excel sheet data
    xl = pd.read_csv(InputFile)

    # # replace '/' in some input fields (like org name) which cause issues when used as directories
    xl['org_name'] = xl['org_name'].str.replace("/","-").str.replace("'","-").str.replace(",","-")
    xl = xl[~xl['org_name'].isin(['ROV Test Silo','ROV Training Silo','ROV Sample Silo'])]  # delete test orgs

    # xl['team_name'] = xl['team_name'].str.replace("/","-").str.replace("'","-").str.replace(",","-")

    # # create a date field as a datetime object
    # xl['date2'] = pd.to_datetime(xl['created_at'])
    minDate2 = xl['created_at'].min()
    maxDate2 = xl['created_at'].max()

    xl['small'] = np.where( xl['addresses_count'] < smallLimit, 1, 0 )
    xl['medium'] = np.where( xl['addresses_count'] < medlimit, 1, 0 )

    xl = xl.rename(columns={'request_id' : 'total_requests'})

    # df = df.rename(columns={'oldName1': 'newName1', 'oldName2': 'newName2'})

    # # xl['date2'] = pd.to_datetime("2021/02/01")
    #
    # # 'daysoffset' will contain the weekday 'day of week' as integers so we can step back to Monday
    # xl['daysoffset'] = xl['date2'].apply(lambda x: x.weekday())
    # # We apply, row by row (axis=1) a timedelta operation
    # xl['data_week_ending'] = xl.apply(lambda x: x['date2'] - dt.timedelta(days=x['daysoffset']-6), axis=1).dt.date
    # xl['month'] = pd.to_datetime(xl['date2']).dt.to_period('M')
    #
    # xl['team2']=xl['team_name'] # Need two copies of team variable for pivots - one for row, one for cols # FIXME: I believe this was removed and two tables used
    #
    # # xl.fillna(" No Team", inplace = True) # note space in front for sorting # change to only replace one col
    # # xl[c("team_name")][ is.na(  xl[c("team_name")]  ) ] = " No Team"
    # # ttt = xl["team_name"].isnull() = " No Team"  # Fixme: THIS NEEDS FIXING
    # xl['team_name'] = np.where(xl['team_name'].isnull() == True, " No Team", xl['team_name'] )

    # teams = xl['team_name'].unique()
    # teams.sort()

    # state_list = ["AL", "AK", "AS", "AZ", "AR", "CA", "CO", "CT", "DE", "DC", "FL", "GA", "GU", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "MP", "OH", "OK", "OR", "PA", "PR", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "VI", "WA", "WV", "WI", "WY"]

    # xl['master_campaign'] = xl['parent_campaign_name'].apply(lambda x: x[0:2])  # changed 4/25/22 BEK to use factory or state
    # xl['master_campaign'] = np.where(xl['factory_name'].isnull() == False, xl['factory_name'], xl['parent_campaign_name'].apply(lambda x: x[0:2]) ) # BEK 4/26/22

    #   4/26/22 if factory name exists, trim parent campaign name  to just state/county
    # xl['parent_campaign_name'] =  np.where(xl['factory_name'].isnull() == True, xl['parent_campaign_name'],
    #                              xl['parent_campaign_name'].str.split("-").str[0]+ '-'+ xl['parent_campaign_name'].str.split("-").str[1] )
    #
    # master_campaigns = xl['master_campaign'].unique()
    # # delete those that are not states (some campaign names ended up 'do not use VA-Loudon...' etc  # FIXME: we may want to include these even though state is bad; many werre written in "DO not use VA-xxx"
    # master_campaigns = [x for x in master_campaigns if x in state_list]
    # master_campaigns.sort()

    ### Create a sheet w pivot tabs for all orgs in separate directory
    # print("Org: Enterprise wide")


    # Run standalone pivot on campaigns
    # df_pt = pd.pivot_table(xlo, columns=['data_week_of'], index=['parent_campaign_name', 'team_name'], values=['addresses_count'], margins=True, dropna=True,aggfunc=np.sum)
    # df_pt = pd.pivot_table(xl, index=['org_name'],
    #                        values=[ 'small', 'medium'], aggfunc='sum',
    #                        margins=True, dropna=True )

    df_pt = pd.pivot_table(xl, index=['org_name'],
                           values=[ 'addresses_count','total_requests', 'small', 'medium'],
                           aggfunc={'addresses_count': 'sum',
                                    'total_requests': 'count',
                                    'small': 'sum',
                                   'medium': 'sum'},
                           margins=True, dropna=True )
    df_pt['pct_small'] = round((df_pt['small'] / df_pt['total_requests'] ) * 100,1)
    df_pt['pct_med'] = round((df_pt['medium'] / df_pt['total_requests'] ) * 100,1)
    df_pt.sort_values(['total_requests'], ascending=[False], inplace=True)
    # df.sort(['a', 'b'], ascending=[True, False])
    df_pt = df_pt[["addresses_count", "total_requests", "small", "medium","pct_small", "pct_med"]]

    # aggfunc = {'D': np.mean,
    #            'E': [min, max, np.mean]}

    # # Use Pandas dataframe writer here because it supports pivot tables and openpyxl does not
    file = "ROVWide Room Summary Showing Small Writers.xlsx"
    if os.path.exists(file):
        os.remove(file)

    # writer = pd.ExcelWriter(file)
    # df_pt.to_excel(writer, sheet_name='Rooms', startrow=6)  # summary sheet of al teams combined
    # writer.save()

    # Use openpyxl here to reference and change individual cells for titles

    # df_pt.columns = df_pt.columns.to_flat_index()

    wb = openpyxl.Workbook()

    # ws_write = wb.create_sheet(0)
    ws_write = wb.worksheets[0]
    # openpyxl.utils.dataframe.dataframe_to_rows(df_pt, index=True, header=True)
    for r in dataframe_to_rows(df_pt, index=True, header=False):
        ws_write.append(r)

    for index, header in enumerate(df_pt.columns.values,start=2): # start 2 because col 1 is index
        ws_write.cell(column=index, row=1, value=header)

    ws_write.insert_rows(1, amount=5)
    wb.save(file)

    for i in range(row):
        ws_write.append([datalist[i][0]])

    wb = openpyxl.load_workbook(file)
    for row in df_pt:
        for i in range(row):
            for j in range(col):
                ws_cell(row=i, col=j).value = datalist[i][j]
    # wb.save(filename='data.xlsx')


    for sh in wb.worksheets:
        autosizexls(sh)

    for sh in wb.worksheets:
        sh['A1'].value = "ROV Address Request by Room Showing Request Size"
        sh['A1'].font = Font(b=True, size=20)
        sh['A3'].value = "Small are < " + str(smallLimit) + ", Medium < " + str(medlimit)
        sh['A3'].font = Font(b=True, size=12)
        sh['A4'].value = "Date range, inclusive: " + minDate2 + " to " + maxDate2
        sh['A4'].font = Font(size=12)
        sh['A5'].value = "Source: " + VLdatafile
        sh['A5'].font = Font(size=12)

    wb.save(file)
    # Done with standalone enterprise wide report

    a=1

if __name__ == '__main__':
    smallWriterReport()

a=1
