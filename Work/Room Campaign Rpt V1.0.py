# Read csv exported from VoterLetters and report by campaign in each room, specifcally seeing if a state has been assigned

# Google cloud console for authorizations: https://console.cloud.google.com/home/dashboard?project=voterletters-reports,
# then 'APi and Services' (upper left)>credentials.  +Create > Oauth > desktop
# delete token.json / credentials.json.  Rename to credentials, copy in

# TODO: Move off my personal google account
# TODO: Can we make TK go to front so it doesn't get missed / crash?
# TODO: set autosize to limit col width; wrap if greater
# TODO: auto set permissions to Barbara and email when uploaded

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from datetime import datetime
import datetime as dt
import xlsxwriter
import pymsgbox
import os
from pathlib import *
import pathlib
import glob
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from random import random
import numpy as np
import pandas as pd
from tkinter import *
from tkinter import Tk  # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory
from tkinter import messagebox
import ast
import google.auth.transport.requests
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from apiclient.http import MediaFileUpload
# from apiclient.discovery import build

upload_notification_list = ['comstockrov@gmail.com']
# upload_notification_list = ['kramsman@yahoo.com','briank@kramericore.com']
upload_notification_msg = "Barbara - This is a new version of the 'Campaign Not Yet written' report.  Relax.  I'm reaching out to Organizers - this is just FYI and reading enjoyment!"

# for google drive api
SCOPES = ['https://www.googleapis.com/auth/drive']
ORG_REPORT_FOLDER_ID = "1OEAdTnhQoAKpyzFqYdM1ztr42vHAqbk5"  # Folder holding VoterLetters Organizer Reports
ADMIN_REPORT_FOLDER_ID ='1_vTXsMYOwK_TLYaRqC7YA_jolBuqDWP4'

sendPermissionAddEmail = True  # send permission granted emails

noteLines = [
    '',
    'WHAT IS IN THIS REPORT',
    '',
    'What you’ll get:',
    '   - Each spreadsheet presents the address and request counts in your room: by time-period,'
    'state, campaign, team, and writer.',
    '   - The tabs (located at the bottom of the sheet) provide an overview then gets more specific:',
    '       the first few tabs show data for your whole room.  Tabs after that are one for each team.',
    '   - A weekly version showing counts by week will be produced each Monday. Monthly time periods',
    '       will be shown in a report produced on the first day of each month.',
    '   - You will still get the daily summary reports from VoterLetters.',
    '',
    'The overall report:',
    '   - The sheet shows only the address and request counts for your room.',
    '   - Each tab shows totals over time, either in weekly or monthly increments.  Dates are endings.',
    '   - The titles give the',
    '         - name of your room',
    '         - name of the team',
    '         - time increment (weekly or monthly)',
    '         - dates included in the report',
    '   - The data will be summarized in each row',
    '   - The reports show total state, campaign, teams, and writers',
    '',
    'The individual tabs of the report are:',
    '   - Tab 1, named  “Notes” explains the report'
    '   - Tab 2, named  “Campaigns” shows requested addresses broken by Campaign',
    '   - Tab 3, “All Team Sums” by Teams',
    '   - Tab 4, “Teams w Writers” by Writers within Teams',
    '   - Tab 5, “Teams w Counts” number of requests by Writers within Teams. ',
    '   - The tabs following are specific to each team (including individual writers).',
    '       These can be copied and pasted for team leads.  '
]

def autosizexls(ws):
    # CHANGED BEK3/10/22 col max width set to 60.
    # TODO: Make col max width an optional parameter; make wrap column if over max an option
    # from https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    # print(ws.title)
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                # print("cell value", cell.value, "data_type",cell.data_type)
                if(cell.data_type == 'd'):
                    dateWidth = 10
                else: dateWidth = len(str(cell.value))
                # dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                # dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), dateWidth))
                dims[cell.column_letter] = (  60 if max((dims.get(cell.column_letter, 0), dateWidth)) >60 else max((dims.get(cell.column_letter, 0), dateWidth))   )

    for col, value in dims.items():
        # print(col , value)
        ws.column_dimensions[col].width = value + 1

def getFolderIds(foldernm, par, drive):
    # added drive parameter
    q1 = "mimeType='application/vnd.google-apps.folder' and trashed=false and name='" + foldernm + "' and parents = '" + par +"'"
    idList = []
    page_token = None
    while True:
        response = drive.files().list(
            q=q1,
            spaces='drive',
            fields="nextPageToken, files(id, name, trashed, parents)",
            pageToken=page_token).execute()
        print("Found old folders with same name: ",response)
        for file in response.get('files', []):
            idList.append(file.get('id'))
            print('Found folder (name,id,trashed,parent):', (file.get('name'), file.get('id'), file.get('trashed'), file.get('parents')))

        page_token = response.get('nextPageToken', None)
        if page_token is None:
            break
    return idList

def getFileIds(filenm, par, drive):  # NOTE this was altered to include drive which wasn't in scope when called from a routine
    q1 = "mimeType != 'application/vnd.google-apps.folder' and trashed=false and name='" + filenm + "' and parents = '" + par +"'"
    idList = []
    page_token = None
    while True:
        response = drive.files().list(
            q=q1,
            spaces='drive',
            fields="nextPageToken, files(id, name, trashed, parents)",
            pageToken=page_token).execute()
        print(" Found files: ", response)
        for file in response.get('files', []):
            idList.append(file.get('id'))
            print('Found file (name,id,trashed,parent):', (file.get('name'), file.get('id'), file.get('trashed'), file.get('parents')))

        page_token = response.get('nextPageToken', None)
        if page_token is None:
            break
    return idList

def getCreds():
    # get credentials needed to set up API service.  Was inline code; this cleaned things up.
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return creds

def Main_Code():
    #### Start Main code
    ####
    # Identify which VoterLetters files should be downloaded before starting
    if True:
        pymsgbox.alert("Download data from VoterLetters before running.\n\n"\
                       "1. Addresses assigned (to-date TOMORROW to get today included):\n      VoterLetters All Enterprise > ROV >\n      Reports >\n      New Report >\n      All Parent Campaign Address Requests.\n  Dates 1/1/22 to prior Monday INCLUSIVE (includes to the day prior to specified).\n\n"\
                       "2. Users (to assign google read permissions):\n      VoterLetters Child Organizations >\n      Export Users to CSV.\n",\
                       "Get ready!")

    choice = pymsgbox.confirm("Run Reports or upload files to Google Drive", 'Run, Upload or Exit?',
                              ["Run", 'Upload', 'Exit'])
    # choice = 'Run'
    if choice == "Exit":
        pymsgbox.alert("Exiting", "Alert")
        exit()
    elif choice == "Upload":
        # pymsgbox.alert("** Uploading Here  **", "Alert")
        # if True:
        #     if True:
        #         Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
        #         inputDir = os.path.expanduser("/Users/Denise/Downloads/")
        #         InputFile = askopenfilename(initialdir=os.path.expanduser(inputDir),
        #                     title="Select VoterLetters export file for USER LIST (enterprise-users-yyyy-mm-dd.csv)",
        #                     filetypes=(("CSV files",
        #                                 "*.csv"),))  # show an "Open" dialog box and return the path to the selected file
        #         df = pd.read_csv(InputFile,
        #                          usecols=['name', 'email', 'role', 'organization'])
        #     else:
        #         df = pd.read_csv("/Users/Denise/Downloads/enterprise-users-2022-01-21.csv",
        #                          usecols=['name', 'email', 'role', 'organization'])
        #     df = df.sort_values(["organization", "name"], ascending=(True, True))
        #     orgEmailList = df.loc[df['role'].isin(['organizer']), ["organization", "email"]].values.tolist()
        # else:
        #     orgEmailList = [
        #         ['NY-NYC and Long Island', 'kramsman+nycorg@gmail.com'],
        #         ['NY-NYC and Long Island', 'meika@centerforcommonground.org'],
        #         ['CA-Peninsula-South Bay', 'dee@centerforcommonground.org'],
        #         ['Mountain States', 'comstockrov@gmail.com'],
        #         ['CA-Peninsula-South Bay', 'nancy@centerforcommonground.org'],
        #     ]

        # TODO: for special gmail emails, kyoko+org@gmail.com, remove everything past the + for permission.

        creds = getCreds() # function call to clean up  # TODO: catch expired credentials and pymsgbox error

        try:
            drive_service = build('drive', 'v3',
                    credentials=creds)  # copied from https://stackoverflow.com/questions/11472401/looking-for-example-using-mediafileupload

            # Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
            # dirStartLoc = os.path.expanduser("~/Dropbox/Postcard Files/VL Org Reports")
            # strReportDirToUpload = askdirectory(initialdir=dirStartLoc, title="Select report directory to UPLOAD")
            # # strReportDirToUpload = '/Users/Denise/Dropbox/Postcard Files/VL Org Reports/all-parent-campaigns-requests-2022-01-26-W'
            # if strReportDirToUpload == "":
            #     pymsgbox.alert("No directory chosen - exiting", "Alert")
            #     exit()
            # reportDirToUpload = pathlib.Path(strReportDirToUpload)
            # folderName = reportDirToUpload.parts[-1]  # the last section of the path
            # reportBy = strReportDirToUpload[-1]
            #
            # # id list of folders to trash(delete)
            # deleteList = getFolderIds(folderName, ORG_REPORT_FOLDER_ID)
            # for folderId in deleteList:
            #     drive_service.files().update(fileId=folderId, body={'trashed':True}).execute()  # trash instead of delete moves to trash; delete gone for good
            # a=1
            #
            # # Create sub folder
            # file_metadata = {
            #         'name': folderName,
            #         'parents': [ORG_REPORT_FOLDER_ID],
            #         'mimeType': 'application/vnd.google-apps.folder' }
            # file = drive_service.files().create(body=file_metadata,
            #         fields='id').execute()
            # newFolderId = file.get('id')
            # print('Folder ID created: %s' % file.get('id'))
            #
            # # get list of filenames in computer directory
            # excelReportFiles = glob.glob(str(reportDirToUpload / '[!~]*.xlsx'))  # xlsx files not starting with ~ (temporary files)
            # for fileWPath in excelReportFiles:
            #     # filePath = pathlib.Path('/Users/Denise/Dropbox/Postcard Files/VoterLetters Reports/all-parent-campaigns-requests-2022-01-20-W')
            #     # path, fileNameWOPath = os.path.split(fileWPath)
            #     fileName = os.path.basename(fileWPath)  # returns filename with extension
            #     fileNameWOExt = pathlib.Path(fileWPath).stem  # filename without extension
            #     roomName = fileNameWOExt[0:-34]  # lop off the additional 34 chars of info added to room for filename like date
            #     print('\nUploading filename, RoomName: ', fileName, ", ", roomName)
            #
            #     file_metadata = {'name': fileNameWOExt,
            #             'mimeType': 'application/vnd.google-apps.spreadsheet',
            #             "parents": [newFolderId]}
            #     media = MediaFileUpload(fileWPath,
            #             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')  # converts xlsx
            #     file = drive_service.files().create(body=file_metadata, supportsAllDrives=True,
            #             media_body=media,
            #             fields='id').execute()
            #     # print('Uploaded file ID: %s' % file.get('id'))
            #     uploadedFileId = file.get('id')
            #
            #     for org, email in orgEmailList:
            #         # email = 'briank@kramericore.com'
            #         if org.replace(",","-") == roomName:  # had to replace , with - in file names so this makes room match.
            #             print('  - Adding permission for: roomName, email: ', roomName, email)
            #             payload = {
            #                 "role": "reader",
            #                 "type": "user",
            #                 "emailAddress": email
            #             }
            #             if sendPermissionAddEmail:
            #                 # TODO: need to trap adding permission when notification off because people without google accounts error.
            #                 if fileNameWOExt[-2:].lower() == '-m':
            #                     emailMessageMsg = "A new MONTHLY VoterLetters summary report is available for your room.  Click to open."
            #                 else:
            #                     emailMessageMsg = "A new WEEKLY VoterLetters summary report is available for your room.  Click to open."
            #                 drive_service.permissions().create(fileId= uploadedFileId, sendNotificationEmail = True, emailMessage = emailMessageMsg, body=payload).execute()
            #             else:
            #                 # Same as above .create command but do not include email text as not allowed if no email sent
            #                 drive_service.permissions().create(fileId= uploadedFileId, sendNotificationEmail = False, body=payload).execute()

            # Upload single Admin report
            Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
            dirStartLoc = os.path.expanduser("~/Dropbox/Postcard Files/VL Admin Reports")
            strAdminReportToUpload = askopenfilename(initialdir=dirStartLoc,
                    title="Select ADMIN report to UPLOAD",
                    filetypes=(("Xlsx files", "*.xlsx "),))
            if strAdminReportToUpload != "":
                # do the upload
                adminReportToUpload = pathlib.Path(strAdminReportToUpload)
                fileName = os.path.basename(adminReportToUpload)  # returns filename with extension
                fileNameWOExt = pathlib.Path(adminReportToUpload).stem  # filename without extension

                deleteList = getFileIds(fileNameWOExt, ADMIN_REPORT_FOLDER_ID, drive_service)
                for fileId in deleteList:
                    drive_service.files().update(fileId=fileId, body={'trashed':True}).execute()  # trash instead of delete moves to trash; delete gone for good
                print('Uploading filename ', fileNameWOExt)

                file_metadata = {'name': fileNameWOExt,
                        'mimeType': 'application/vnd.google-apps.spreadsheet',
                        "parents": [ADMIN_REPORT_FOLDER_ID]}
                media = MediaFileUpload(adminReportToUpload,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')  # converts xlsx
                file = drive_service.files().create(body=file_metadata, supportsAllDrives=True,
                        media_body=media,
                        fields='id').execute()
                uploadedFileId = file.get('id')

                print("Adding permission for selected few")
                choice = pymsgbox.confirm("Send to emails with msg?\n\nEmails:\n" + "\n".join(upload_notification_list) +"\n\nMessage:\n" + upload_notification_msg, 'Continue or Exit?',
                                          ["Continue", "Exit"])
                if choice == "Exit":
                    pymsgbox.alert("Exiting", "Alert")
                    exit()

                for email in upload_notification_list:
                    payload = {
                        "role": "reader",
                        "type": "user",
                        "emailAddress": email
                    }
                    # if sendPermissionAddEmail:
                    drive_service.permissions().create(fileId= uploadedFileId, sendNotificationEmail = True, emailMessage = upload_notification_msg, body=payload).execute()

                # print('Uploaded file ID: %s' % file.get('id'))

        except HttpError as err:
            print(err)

    elif choice == "Run":
        # pymsgbox.alert("** Running Here  **", "Alert")

        # reportBy = pymsgbox.confirm('[W]eekly or [M]onthly report?', 'Date Format', ["W", "M", 'Cancel'])
        # if reportBy.lower() == 'm':
        #     reportVar = "month"
        # elif reportBy.lower() == 'w':
        #     reportVar = "data_week_ending"
        # else:
        #     exit()

        # Put enterprise wide reports in a separate directory
        # outputDirAdmin = os.path.expanduser("~/Dropbox/Postcard Files/Other/VoterLetters/Reports/VL ROVWide Reports")
        outputDirAdmin = os.path.expanduser("~/Dropbox/Postcard Files/VL Admin Reports")

        # Dir to prompt to start looking for input file
        inputDir = os.path.expanduser("/Users/Denise/Downloads/")
        # Base of report file; input file name is used to create a sub dir under this
        outputDir = os.path.expanduser("~/Dropbox/Postcard Files/VL Org Reports")

        Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing

        # InputFile = askopenfilename(initialdir=os.path.expanduser(r"/Users/Denise/Downloads/"),title="Select VoterLetters address export file", filetypes=(("CSV files", "*.csv "),))  # show an "Open" dialog box and return the path to the selected file
        InputFile = askopenfilename(initialdir=os.path.expanduser(inputDir),title="Select VoterLetters address export file (all-parent-campaigns-requests-yyyy-mm-dd.csv)", filetypes=(("CSV files", "*.csv "),))  # show an "Open" dialog box and return the path to the selected file
        # tried but failed InputFile = askopenfilename(initialdir=os.path.expanduser(inputDir),title="Select VoterLetters address export file", filetypes=('VL files','all-parent*.*'))  # show an "Open" dialog box and return the path to the selected file
        # TODO: set filter to only display fileNAMES that are proper input, such as 'all-parent*.csv'

        # Get file containing Writer info
        writerFile = askopenfilename(initialdir=os.path.expanduser(inputDir),
                                    title="Select VoterLetters export file for USER LIST (enterprise-users-yyyy-mm-dd.csv)",
                                    filetypes=(("CSV files", "*.csv"),))

        # head, tail = os.path.split(InputFile)
        VLdatafile = os.path.basename(InputFile)
        # outputDirWFile = os.path.join(outputDir,Path(InputFile).stem + "-" + reportBy)
        outputDirWFile = os.path.join(outputDir,Path(InputFile).stem )
        filedate = Path(InputFile).stem[-10:]

        # # CHECK and create dir if not exists
        # if not os.path.isdir(outputDirWFile):
        #     # pymsgbox.alert("Input Path does not exist. Creating.\n\n" + outputDirWFile, "Fixing Directory Structure")
        #     # FIXME: close Tkinter window here
        #     os.makedirs(outputDirWFile)

        # Create dataframe of excel sheet data
        requestData = pd.read_csv(InputFile)

        # replace '/' in some input fields (like org name) which cause issues when used as directories
        requestData['org_name'] = requestData['org_name'].str.replace("/", "-").str.replace("'", "-").str.replace(",", "-")
        requestData['team_name'] = requestData['team_name'].str.replace("/", "-").str.replace("'", "-").str.replace(",", "-")
        requestData['master_campaign'] = requestData['parent_campaign_name'].apply(lambda x: x[0:2].upper())

        # create a date field as a datetime object
        requestData['date2'] = pd.to_datetime(requestData['created_at'])
        minDate2 = requestData['created_at'].min()
        maxDate2 = requestData['created_at'].max()
        # xl['date2'] = pd.to_datetime("2021/02/01")


        class room:
            def __init__(self, room_name, total_address_count):  # , campaigns=[], master_campaigns=[]):
                self.name = room_name
                self.total_address_count = total_address_count
                self.master_campaigns = []
                self.campaigns = []
                self.organizerEmails = []
                self.organizerNames = []


        # get total counts by room
        df_roomCounts = pd.pivot_table(requestData, index=['org_name'],
                               values=['addresses_count'], dropna=True, aggfunc=np.sum)
        df_roomCounts = df_roomCounts.reset_index()  # converts multi-index to columns
        # df_roomCounts = df_roomCounts[df_roomCounts['org_name'] not in ['ROV Training Silo']]
        df_roomCounts.drop(df_roomCounts[df_roomCounts.org_name == 'ROV Training Silo'].index, inplace=True)

        #         d.rename(columns={'address': 'address_count'}, inplace=True)


        rooms = []

        for index, row in df_roomCounts.iterrows():
            rooms.append(room(row.org_name, row.addresses_count))

        # counts by room by master campaign
        df_masterCampaignCounts = pd.pivot_table(requestData, index=['org_name','master_campaign'],
                               values=['addresses_count'], dropna=True, aggfunc=np.sum)
        df_masterCampaignCounts = df_masterCampaignCounts.reset_index()  # converts multi-index to columns

        orgs_unique = df_masterCampaignCounts['org_name'].unique()
        masterList  = [ (org1, [row[1] for index, row in df_masterCampaignCounts.iterrows() if row[0] == org1 ]) for org1 in orgs_unique ]

        # turn list into a dictionary
        dictMastCamps = {org: master for org, master in masterList}

        for room in rooms:
            room.master_campaigns = dictMastCamps.get(room.name,'no masters')

        # counts by room by campaign
        df_campaignCounts = pd.pivot_table(requestData, index=['org_name','parent_campaign_name'],
                               values=['addresses_count'], margins=True, dropna=True, aggfunc=np.sum)
        df_campaignCounts = df_campaignCounts.reset_index()  # converts multi-index to columns
        campaignList  = [ (org1, [ [row[1],row[2]] for index, row in df_masterCampaignCounts.iterrows() if row[0] == org1 ]) for org1 in orgs_unique ]

        dictCamps = {org: camps for org, camps in campaignList}

        for room in rooms:
            room.campaigns = dictCamps.get(room.name,'no campaigns')

        # Get file containing Writers
        # moved up top along with select Request file so no lag while calculations taking place
        # writerFile = askopenfilename(initialdir=os.path.expanduser(inputDir),
        #                             title="Select VoterLetters export file for USER LIST (enterprise-users-yyyy-mm-dd.csv)",
        #                             filetypes=(("CSV files", "*.csv"),))
        writerData = pd.read_csv(writerFile, usecols=['name', 'email', 'role', 'organization'])
        # replace '/' in some input fields (like org name) which cause issues when used as directories
        writerData['organization'] = writerData['organization'].str.replace("/", "-").str.replace("'", "-").str.replace(",", "-")
        # writerData = writerData[writerData['role'] == 'organizer']

        # Get organizer emails
        df_organizerEmails = pd.pivot_table(writerData[writerData['role'] == 'organizer'], index=['organization','email'],
                             dropna=True, aggfunc='count')
                               # values=['addresses_count'], margins=True, dropna=True, aggfunc=np.sum)
        df_organizerEmails = df_organizerEmails.reset_index()  # converts multi-index to columns
        # df_organizers = df_organizers.drop(['index'], axis=1)
        organizerEmailList  = [ (org1, [ row[1] for index, row in df_organizerEmails.iterrows() if row[0] == org1 ]) for org1 in orgs_unique ]

        dictOrganizerEmails = {org: emails for org, emails in organizerEmailList}

        for room in rooms:
            room.organizerEmails = dictOrganizerEmails.get(room.name, 'no organizers')


        # Get organizer names
        df_organizerNames = pd.pivot_table(writerData[writerData['role'] == 'organizer'], index=['organization','name'],
                             dropna=True, aggfunc='count')
        df_organizerNames = df_organizerNames.reset_index()  # converts multi-index to columns
        organizerNameList  = [ (org1, [ row[1] for index, row in df_organizerNames.iterrows() if row[0] == org1 ]) for org1 in orgs_unique ]

        dictOrganizerNames = {org: names for org, names in organizerNameList}

        for room in rooms:
            room.organizerNames = dictOrganizerNames.get(room.name, 'no organizers')


        #### Room information is filled.  Report out #####
        state_list = set(["AL", "AK", "AS", "AZ", "AR", "CA", "CO", "CT", "DE", "DC", "FL", "GA", "GU", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "MP", "OH", "OK", "OR", "PA", "PR", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "VI", "WA", "WV", "WI", "WY"])

        # set output sheet for report
        file = os.path.join(os.path.expanduser(outputDirAdmin), "VoterLetters Room-Campaign Report " + filedate + ".xlsx")
        writer = pd.ExcelWriter(file)

        # check_mast = input('Input the master campaign')
        # master_campaigns = requestData['master_campaign'].unique()  # TODO: use only those found in state list
        master_campaigns_set = set(requestData['master_campaign'])
        master_campaigns = master_campaigns_set.intersection(state_list)

        for check_mast in master_campaigns:

            missing_master = [( room.name, room.total_address_count, ", ".join(room.organizerEmails), ", ".join(room.organizerNames) ) for room in rooms if check_mast not in room.master_campaigns]

            df = pd.DataFrame(missing_master)
            df = df.rename(columns={df.columns[0]: 'Room Name', df.columns[1]: 'Address Count All Campaigns', df.columns[2]: 'Organizer Emails',df.columns[3]: 'Organizer Names'})
            # df.sort_values(['Address Cnt All Campaigns', 'Room Name'], ascending=[False,False], inplace=True)
            df.sort_values(['Room Name'], ascending=[True], inplace=True)

            # print('Missing ' + check_mast+ ' campaign\n')
            # print('\n'.join(map(str,missing_master)))

            df.to_excel(writer, sheet_name=check_mast, startrow=6, index=False)
        writer.save()

            # Use openpyxl here to reference and change individual cells for titles
        wb = openpyxl.load_workbook(file)

        for sh in wb.worksheets:
            autosizexls(sh)

        for sh in wb.worksheets:
            sh['A1'].value = "Rooms Having Not Yet Written A Master Campaign"
            sh['A1'].font = Font(b=True, size=20)
            if sh.title not in ['Campaigns','Rooms']:
                sh['A2'].value = "Master Campaign: " + sh.title
            sh['A2'].font = Font(b=True, size=16)

            sh['A4'].value = "Date range, inclusive: " + minDate2 + " to " + maxDate2
            sh['A4'].font = Font(size=12)
            sh['A5'].value = "Source: " + VLdatafile
            sh['A5'].font = Font(size=12)

        wb.save(file)

        print("\nDone with all masters.")
        pymsgbox.alert("Org reports produced. In:\n\n" + outputDir + "\n\n\nAdmin reports produced. In:\n\n" + outputDirAdmin, "Done!")
    ### End of main_code

Main_Code()
